#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EdgeProfiler 2.0 - Automated HIL Testing & Multi-Node Monitoring Tool
=====================================================================
Industry-grade QA dashboard for the ESP32-S3 LoRa edge system
(SHT30 + BH1750 + BMP180 nodes  ->  LoRa gateway, AES/ASCON encrypted).

Highlights
----------
* MULTI-PORT  : connect to several COM ports at once (gateway + node consoles).
* MULTI-NODE  : auto-discovers nodes from gateway telemetry, per-node plots/tiles.
* GATEWAY VIEW: heap, dual-core CPU, uptime, nodes online, WiFi, Firebase status.
* RICH METRICS: response time (RTT), distance, PDR, proc/enc time, bandwidth(pkt/h),
                packet loss, anomaly-detection accuracy, gateway-loss resilience.
* CLICKABLE TILES -> detail pop-ups with history + min/max/avg.
* CORRELATION : distance vs packet-loss scatter (per node).
* TEST SUITE  : fault injection, jamming, FSM override, gateway-offline, bandwidth,
                encryption verify, multi-node stress, full auto-run.
* FREEZE/RESUME live plotting, HTML + XML reports.

Telemetry the gateway emits (one JSON object per line over USB serial):
  {"type":"node","node":1,"sys":{"heap":..,"c0_cpu":..,"c1_cpu":..},
   "sensor":{"t":..,"h":..,"lux":..,"press":..,"bmp_t":..},
   "lora":{"rssi":..,"snr":..,"pdr":..,"ul_rssi":..},
   "metrics":{"rtt":..,"dist":..,"proc":..,"enc":..,"resp":..,"gwproc":..,"pph":..,"algo":".."},
   "alert":0,"fsm":"SAFE","online":1}
  {"type":"gw","sys":{..},"uptime":..,"nodes_online":..,"wifi":1,"fb_ok":1,"algo":".."}
"""

import sys
import json
import time
import math
import threading
import datetime
import statistics
from collections import deque, defaultdict

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
import pyqtgraph as pg
import serial
import serial.tools.list_ports
from jinja2 import Template
import xml.etree.ElementTree as ET

# --------------------------------------------------------------------------
APP_NAME = "EdgeProfiler"
APP_VERSION = "2.1"   # +OTA Update Test, +LoRa Replay Attack test
HIST = 400                 # samples kept per series
CORR_MAX = 600             # correlation points kept
DEFAULT_BAUD = 115200
RECORD_BYTES = 64          # RAM size of one offline record on the node (epoch+payload)

C = {  # palette
    "t": "#ff5c5c", "h": "#4fc3f7", "press": "#ffd166", "lux": "#a0e060",
    "rssi": "#ff8a65", "snr": "#ba68c8", "rtt": "#00e5ff", "dist": "#ffca28",
    "pdr": "#26c281", "proc": "#f48fb1", "enc": "#9ccc65", "heap": "#26c281",
    "pph": "#80d8ff",
}
BG_NORMAL = "#161b22"
BG_ALERT = "#3a1414"
NODE_COLORS = ["#4fc3f7", "#ff8a65", "#a0e060", "#ba68c8", "#ffd166",
               "#26c281", "#f06292", "#80cbc4"]

# Log-distance path-loss model used to convert RSSI -> distance (calibratable
# live from the GUI, no reflash needed):
#     RSSI(d) = RSSI_REF_1M - 10*n*log10(d)   =>   d = 10^((ref - rssi)/(10n))
# The defaults are deliberately tuned for short range; USE THE CALIBRATE BUTTON
# with a known distance to make it accurate for your hardware/antenna.
CALIB = {"ref": -49.0, "n": 2.0}     # ref = RSSI(dBm) measured at exactly 1 m


def rssi_to_distance(rssi):
    try:
        if rssi == 0:
            return 0.0
        d = 10.0 ** ((CALIB["ref"] - float(rssi)) / (10.0 * CALIB["n"]))
        return max(0.0, min(d, 9999.0))
    except Exception:                                  # noqa: BLE001
        return 0.0


# ==========================================================================
# Serial worker (one per COM port)
# ==========================================================================
class SerialWorker(QThread):
    node_tel = pyqtSignal(str, dict)     # (port, telemetry dict)  type=node
    gw_tel = pyqtSignal(str, dict)       # (port, telemetry dict)  type=gw
    tasks = pyqtSignal(str, dict)        # (port, FreeRTOS task list) type=node_tasks
    ota = pyqtSignal(str, dict)          # (port, OTA progress)     type=ota
    sec = pyqtSignal(str, dict)          # (port, security event)   type=sec
    resp = pyqtSignal(str, dict)         # (port, command response)
    raw = pyqtSignal(str, str)           # (port, raw line)
    status = pyqtSignal(str, bool, str)  # (port, connected, msg)

    def __init__(self, port, baud, parent=None):
        super().__init__(parent)
        self.port = port
        self.baud = int(baud)
        self._ser = None
        self._run = False
        self._lock = threading.Lock()
        self.is_gateway = False          # set true once it emits typed JSON

    def stop(self):
        self._run = False
        self.wait(1500)

    def send_command(self, obj):
        line = json.dumps(obj, separators=(",", ":"))
        with self._lock:
            try:
                if self._ser and self._ser.is_open:
                    self._ser.write((line + "\n").encode())
                    self._ser.flush()
                    self.raw.emit(self.port, "TX  " + line)
                    return True
            except Exception as exc:                       # noqa: BLE001
                self.status.emit(self.port, False, f"TX error: {exc}")
        return False

    def run(self):
        try:
            self._ser = serial.Serial(self.port, self.baud, timeout=0.2)
        except Exception as exc:                           # noqa: BLE001
            self.status.emit(self.port, False, f"open failed: {exc}")
            return
        self._run = True
        self.status.emit(self.port, True, f"connected @ {self.baud}")
        buf = b""
        while self._run:
            try:
                chunk = self._ser.read(512)
                if chunk:
                    buf += chunk
                    while b"\n" in buf:
                        ln, buf = buf.split(b"\n", 1)
                        self._handle(ln)
                else:
                    self.msleep(5)
            except Exception as exc:                       # noqa: BLE001
                self.status.emit(self.port, False, f"rx error: {exc}")
                break
        try:
            if self._ser and self._ser.is_open:
                self._ser.close()
        except Exception:                                  # noqa: BLE001
            pass
        self.status.emit(self.port, False, "disconnected")

    def _handle(self, raw_bytes):
        try:
            line = raw_bytes.decode("utf-8", "replace").strip()
        except Exception:                                  # noqa: BLE001
            return
        if not line:
            return
        self.raw.emit(self.port, line)
        if not (line.startswith("{") and line.endswith("}")):
            return
        try:
            obj = json.loads(line)
        except (json.JSONDecodeError, ValueError):
            return
        if not isinstance(obj, dict):
            return
        t = obj.get("type")
        if t == "node":
            self.is_gateway = True
            self.node_tel.emit(self.port, obj)
        elif t == "node_tasks":
            self.tasks.emit(self.port, obj)
        elif t == "gw":
            self.is_gateway = True
            self.gw_tel.emit(self.port, obj)
        elif t == "ota":
            self.ota.emit(self.port, obj)
        elif t == "sec":
            self.sec.emit(self.port, obj)
        elif "resp" in obj:
            self.resp.emit(self.port, obj)
        elif "sys" in obj or "sensor" in obj:
            # legacy single-node frame
            self.is_gateway = True
            obj.setdefault("node", 1)
            self.node_tel.emit(self.port, obj)


# ==========================================================================
# Data models
# ==========================================================================
class NodeData:
    SERIES = ["t", "h", "lux", "press", "rssi", "snr", "pdr",
              "rtt", "dist", "proc", "enc", "pph", "heap", "cpu", "awake", "sleep",
              "tf", "hf", "pressf", "luxf"]   # *f = filtered (qua bo loc tren node)

    def __init__(self, node_id):
        self.node_id = node_id
        self.x = deque(maxlen=HIST)
        self.s = {k: deque(maxlen=HIST) for k in self.SERIES}
        self.corr = deque(maxlen=CORR_MAX)   # (dist, loss%)
        self.n = 0
        self.last = {}
        self.fsm = "?"
        self.alert = 0
        self.algo = "?"
        self.online = 0
        self.fw_ver = 0          # phien ban firmware (de theo doi OTA tu xa)
        self.ota_from = 0        # fw_ver truoc khi OTA (gateway cung cap)
        self.ota_cmd_epoch = 0   # luc gateway gui lenh OTA
        self.port = ""
        self.last_seen = time.time()
        self.rx = 0           # packets delivered successfully (node ack_total)
        self.sent = 0         # packets the node attempted to send (node tx_total)
        self.sps = 0.0        # sampling rate (sample/s) computed BY THE NODE
        self._rssi_ema = None # smoothed RSSI for a stable distance estimate
        # --- node CPU + watchdog health ---
        self.cpu0 = 0
        self.cpu1 = 0
        self.wdt_s = 0          # watchdog timeout (s)
        self.reset_reason = "?"
        self.boot = 0           # boot counter (RTC-retained)
        self.wdt_resets = 0     # how many times the watchdog reset the node
        self._last_boot = None
        # --- power / sleep ---
        self.sleep_mode = "?"   # none / light / deep
        self.awake_ms = 0
        self.sleep_ms = 0
        # --- FreeRTOS profiler ---
        self.tasks = []         # [{n,cpu,core,stk,prio,st}]
        self.core0 = 0          # per-core CPU from run-time stats
        self.core1 = 0
        # --- data filter (edge processing) ---
        self.filter_algo = "?"  # None / MovingAvg / EMA / Median / Kalman
        self.frame_times = deque(maxlen=60)   # telemetry arrival times (fallback rate)
        # --- offline store-and-forward monitoring ---
        self.buf_count = 0
        self.buf_cap = 0
        self.buf_stored = 0
        self.buf_flushed = 0
        self.buf_dropped = 0
        self.buf_series = deque(maxlen=HIST)   # buffer occupancy over time
        self.backlog_log = deque(maxlen=500)   # replayed records (epoch + values)
        self.events = deque(maxlen=500)        # store/flush/drop event log
        # report accumulators
        self.acc = defaultdict(list)

    def push(self, tel):
        self.n += 1
        self.x.append(self.n)
        sen = tel.get("sensor", {}); lora = tel.get("lora", {})
        sysd = tel.get("sys", {}); m = tel.get("metrics", {})
        self.online = int(tel.get("online", 1))
        nan = float("nan")

        def num(d, k):
            v = d.get(k, None)
            return float(v) if isinstance(v, (int, float)) else None

        def g0(d, k):
            v = d.get(k, 0.0)
            return float(v) if isinstance(v, (int, float)) else 0.0

        # Link valid only if node reports online AND a numeric RSSI is present.
        rssi_v = num(lora, "rssi")
        link = (self.online == 1) and (rssi_v is not None)
        rssi = rssi_v if link else nan
        rtt_v = num(lora, "rtt")
        if rtt_v is None:
            rtt_v = num(m, "rtt")
        rtt = rtt_v if (link and rtt_v is not None) else nan
        snr_v = num(lora, "snr")
        snr = snr_v if (link and snr_v is not None) else nan
        # Distance is computed in the TOOL from RSSI (calibratable). No link -> N/A.
        # RSSI is noisy (esp. at close range) so smooth it before converting,
        # otherwise the distance jumps around wildly.
        if link:
            self._rssi_ema = rssi if self._rssi_ema is None \
                else 0.6 * self._rssi_ema + 0.4 * rssi
            dist = rssi_to_distance(self._rssi_ema)
        else:
            dist = nan

        # Counters: node uses tx/ack, gateway-relay uses sent/rx.
        self.rx = int(m.get("rx", m.get("ack", self.rx)))
        self.sent = int(m.get("sent", m.get("tx", self.sent)))
        self.sps = float(m.get("sps", self.sps))
        pdr_v = num(lora, "pdr")
        if pdr_v is None:
            pdr_v = (100.0 * self.rx / self.sent) if self.sent else 100.0

        pwr = tel.get("power", {})
        self.sleep_mode = pwr.get("mode", self.sleep_mode)
        self.awake_ms = int(g0(pwr, "awake_ms"))
        self.sleep_ms = int(g0(pwr, "sleep_ms"))
        self.cpu0 = int(pwr.get("cpu", g0(sysd, "c0_cpu")))
        self.cpu1 = int(g0(sysd, "c1_cpu"))
        # filtered values (computed on the node). Fall back to raw if absent.
        filt = tel.get("filt", {})
        self.filter_algo = filt.get("algo", self.filter_algo)
        tf = g0(filt, "t") if "t" in filt else g0(sen, "t")
        hf = g0(filt, "h") if "h" in filt else g0(sen, "h")
        pf = g0(filt, "press") if "press" in filt else g0(sen, "press")
        lf = g0(filt, "lux") if "lux" in filt else g0(sen, "lux")

        vals = {
            "t": g0(sen, "t"), "h": g0(sen, "h"), "lux": g0(sen, "lux"),
            "press": g0(sen, "press"), "rssi": rssi, "snr": snr,
            "pdr": pdr_v, "rtt": rtt, "dist": dist,
            "proc": g0(m, "proc"), "enc": g0(m, "enc"), "pph": g0(m, "pph"),
            "heap": g0(sysd, "heap"), "cpu": float(self.cpu0),
            "awake": float(self.awake_ms), "sleep": float(self.sleep_ms),
            "tf": tf, "hf": hf, "pressf": pf, "luxf": lf,
        }
        for k, v in vals.items():
            self.s[k].append(v)
            self.acc[k].append(v)
        if link:
            self.corr.append((dist, max(0.0, 100.0 - pdr_v)))
        self.frame_times.append(time.time())

        # --- offline store-and-forward ---
        buf = tel.get("buf", {})
        if buf:
            self.buf_cap = int(buf.get("cap", self.buf_cap))
            self.buf_stored = int(buf.get("stored", self.buf_stored))
            self.buf_flushed = int(buf.get("flushed", self.buf_flushed))
            self.buf_dropped = int(buf.get("dropped", self.buf_dropped))
            newc = int(buf.get("count", self.buf_count))
            if newc > self.buf_count:
                self.events.append((time.time(),
                    f"STORE  offline buffer {self.buf_count} -> {newc} (link lost)"))
            elif newc < self.buf_count:
                self.events.append((time.time(),
                    f"FLUSH  sent backlog, buffer {self.buf_count} -> {newc}"))
                if newc == 0:
                    self.events.append((time.time(),
                        "DRAINED  buffer empty — all offline data delivered & freed"))
            self.buf_count = newc
        self.buf_series.append(self.buf_count)

        if int(tel.get("backlog", 0)) == 1:
            ep = int(tel.get("epoch", 0))
            tstr = (datetime.datetime.fromtimestamp(ep).strftime("%Y-%m-%d %H:%M:%S")
                    if ep else "unknown-time")
            self.backlog_log.append((ep, vals["t"], vals["h"], vals["press"], vals["lux"]))
            self.events.append((time.time(),
                f"BACKLOG@{tstr}  T={vals['t']:.1f} H={vals['h']:.0f} "
                f"P={vals['press']:.1f} L={vals['lux']:.0f}"))

        # --- watchdog / reset tracking ---
        wdt = tel.get("wdt", {})
        if wdt:
            self.wdt_s = int(wdt.get("timeout_s", self.wdt_s))
            self.reset_reason = wdt.get("reset", self.reset_reason)
            self.wdt_resets = int(wdt.get("wdt_resets", self.wdt_resets))
            b = int(wdt.get("boot", self.boot))
            if self._last_boot is not None and b > self._last_boot:
                # node rebooted between two telemetry frames
                if self.reset_reason in ("TASK_WDT", "INT_WDT", "WDT"):
                    self.events.append((time.time(),
                        f"WATCHDOG RESET  node rebooted (boot #{b}, reason {self.reset_reason})"))
                elif self.reset_reason != "DEEPSLEEP":
                    # DEEPSLEEP reboots are normal each cycle in deep-sleep mode -> don't spam
                    self.events.append((time.time(),
                        f"REBOOT  node restarted (boot #{b}, reason {self.reset_reason})"))
            self._last_boot = b
            self.boot = b

        self.last = tel
        self.fsm = tel.get("fsm", "?")
        self.alert = int(tel.get("alert", 0))
        self.algo = m.get("algo", self.algo)
        self.online = int(tel.get("online", 1))
        fw = tel.get("fw", {})
        self.fw_ver = int(fw.get("ver", self.fw_ver))
        self.ota_from = int(tel.get("ota_from", self.ota_from))
        self.ota_cmd_epoch = int(tel.get("ota_cmd_epoch", self.ota_cmd_epoch))
        self.last_seen = time.time()

    def frame_rate(self):
        ft = list(self.frame_times)
        if len(ft) < 2:
            return 0.0
        span = ft[-1] - ft[0]
        return (len(ft) - 1) / span if span > 0 else 0.0


class GatewayData:
    def __init__(self):
        self.x = deque(maxlen=HIST)
        self.heap = deque(maxlen=HIST)
        self.c0 = deque(maxlen=HIST)
        self.c1 = deque(maxlen=HIST)
        self.n = 0
        self.last = {}
        self.last_seen = 0.0

    def push(self, tel):
        self.n += 1
        sysd = tel.get("sys", {})
        self.x.append(self.n)
        self.heap.append(float(sysd.get("heap", 0)))
        self.c0.append(float(sysd.get("c0_cpu", 0)))
        self.c1.append(float(sysd.get("c1_cpu", 0)))
        self.last = tel
        self.last_seen = time.time()


# ==========================================================================
# Clickable metric tile
# ==========================================================================
class MetricTile(QtWidgets.QFrame):
    clicked = pyqtSignal(str)

    def __init__(self, key, title, unit="", color="#4fc3f7"):
        super().__init__()
        self.key = key
        self.setObjectName("Tile")
        self.setCursor(Qt.PointingHandCursor)
        lay = QtWidgets.QVBoxLayout(self)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(2)
        self.lbl_title = QtWidgets.QLabel(title)
        self.lbl_title.setStyleSheet("color:#8aa; font-size:11px;")
        self.lbl_val = QtWidgets.QLabel("--")
        self.lbl_val.setStyleSheet(f"color:{color}; font-size:22px; font-weight:700;")
        self.lbl_unit = QtWidgets.QLabel(unit)
        self.lbl_unit.setStyleSheet("color:#667; font-size:10px;")
        lay.addWidget(self.lbl_title)
        lay.addWidget(self.lbl_val)
        lay.addWidget(self.lbl_unit)

    def set_value(self, text):
        self.lbl_val.setText(text)

    def mousePressEvent(self, ev):
        self.clicked.emit(self.key)


# ==========================================================================
# Detail pop-up for a metric
# ==========================================================================
class DetailDialog(QtWidgets.QDialog):
    def __init__(self, node_data, key, title, color, parent=None):
        super().__init__(parent)
        self.nd = node_data
        self.key = key
        self.setWindowTitle(f"{title} - Node {node_data.node_id}")
        self.resize(720, 460)
        lay = QtWidgets.QVBoxLayout(self)

        self.plot = pg.PlotWidget()
        self.plot.setBackground(BG_NORMAL)
        self.plot.showGrid(x=True, y=True, alpha=0.25)
        self.plot.setTitle(title, color="#ddd", size="11pt")
        self.curve = self.plot.plot(pen=pg.mkPen(color, width=2))
        lay.addWidget(self.plot, 1)

        self.stats = QtWidgets.QLabel("")
        self.stats.setStyleSheet("font-size:13px;")
        lay.addWidget(self.stats)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.refresh)
        self.timer.start(400)
        self.refresh()

    def refresh(self):
        xs = list(self.nd.x)
        ys = list(self.nd.s.get(self.key, []))
        m = min(len(xs), len(ys))
        if m:
            self.curve.setData(xs[-m:], ys[-m:])
            seg = ys[-m:]
            self.stats.setText(
                f"last: {seg[-1]:.2f}    min: {min(seg):.2f}    "
                f"max: {max(seg):.2f}    avg: {statistics.mean(seg):.2f}    "
                f"samples: {len(seg)}")

    def closeEvent(self, ev):
        self.timer.stop()
        ev.accept()


# ==========================================================================
# Help / documentation pop-up
# ==========================================================================
HELP_HTML = """
<h2 style="color:#4fc3f7">EdgeProfiler — Hướng dẫn sử dụng &amp; ý nghĩa thông số</h2>
<p>Công cụ giám sát/kiểm thử HIL cho hệ ESP32-S3 LoRa (nhiều Node → 1 Gateway,
gói tin mã hóa AES-128-GCM hoặc ASCON-128).</p>

<h3 style="color:#ffd166">1. Kết nối &amp; nguồn dữ liệu</h3>
<ul>
<li><b>Chỉ cần cắm NODE vào máy</b> là theo dõi được — node TỰ phát telemetry JSON
qua USB của chính nó, <b>không phụ thuộc gateway</b>. Gateway có thể đặt ở xa.</li>
<li>Ô <b>Source</b> để chọn nguồn hiển thị:
  <ul>
  <li><b>Node (direct)</b>: đọc thẳng từ node (khuyên dùng cho đề tài tập trung node).</li>
  <li><b>Gateway (relayed)</b>: xem dữ liệu node mà gateway nhận lại được (nếu cắm gateway).</li>
  <li><b>Both (compare)</b>: so sánh node trực tiếp vs gateway nhận được → kiểm chứng
  gateway nhận ĐÚNG &amp; ĐỦ (hiện ✓/✗ và giá trị hai bên).</li>
  </ul></li>
<li>Bấm <b>+ Connect Port</b> nhiều lần để cắm cả node và gateway cùng lúc.</li>
<li>Khi <b>mất kết nối LoRa</b> (rút LoRa gateway), node vẫn phát telemetry với
<b>online=0</b>; các thông số LoRa (RSSI/SNR/RTT/khoảng cách) hiện <b>"—"</b>, FSM hiện
<b>OFFLINE</b>, và node bắt đầu <b>lưu offline</b> (xem panel store-and-forward).</li>
<li><b>Freeze</b>: tạm dừng cập nhật biểu đồ; bấm lại để chạy tiếp.</li>
</ul>
<p><b>Vai trò gateway:</b> chỉ để xác nhận node gửi được qua LoRa và trả ACK —
minh chứng đường truyền/thiết kế LoRa hoạt động đúng. Mọi kịch bản test đều
đánh giá trên NODE.</p>

<h3 style="color:#ffd166">2. Ý nghĩa các thông số (tiles &amp; biểu đồ)</h3>
<table cellpadding="5">
<tr><td><b style="color:#00e5ff">Response (RTT)</b></td>
<td>Thời gian khứ hồi node→gateway→ACK (ms). Node tự đo từ lúc phát đến lúc
nhận ACK. Phản ánh độ trễ phản ứng của hệ.</td></tr>
<tr><td><b style="color:#ffca28">Distance</b></td>
<td>Khoảng cách ước lượng từ RSSI theo mô hình suy hao
<i>d = 10^((RSSI@1m − RSSI)/(10·n))</i>. <b>Cần hiệu chuẩn</b> bằng nút
<i>Calibrate@here</i> (đặt node ở khoảng cách đã biết, nhập số mét, bấm nút).</td></tr>
<tr><td><b style="color:#26c281">PDR</b></td>
<td>Packet Delivery Ratio = số gói gateway nhận / số gói node gửi (theo seq).
100% = không mất gói.</td></tr>
<tr><td><b style="color:#80d8ff">Bandwidth</b></td>
<td>Thông lượng ước lượng (gói/giờ) dựa trên khoảng cách giữa các gói.</td></tr>
<tr><td><b style="color:#f48fb1">Proc time</b></td>
<td>Thời gian node xử lý 1 chu kỳ (đọc cảm biến + dựng gói) tính bằng µs.</td></tr>
<tr><td><b style="color:#9ccc65">Encrypt time</b></td>
<td>Thời gian mã hóa gói (µs). So sánh AES vs ASCON tại đây.</td></tr>
<tr><td><b style="color:#ff8a65">RSSI / SNR</b></td>
<td>Cường độ &amp; tỉ số tín/tạp của tín hiệu LoRa thu được tại gateway (dBm / dB).</td></tr>
<tr><td><b style="color:#26c281">Node heap</b></td>
<td>RAM trống của node (byte) — theo dõi rò rỉ bộ nhớ.</td></tr>
<tr><td><b style="color:#ff7043">Node CPU</b></td>
<td>% CPU node = <i>thời gian thức / (thức + ngủ)</i>. Khi node ngủ (light/deep) CPU thấp;
khi mất kết nối (retry liên tục, không ngủ) CPU cao — đó là lý do trước đây thấy ~100%.</td></tr>
<tr><td><b style="color:#80cbc4">Sleep</b></td>
<td>Chế độ ngủ THỰC SỰ dùng (none / <b>light</b> / <b>deep</b>) + thời gian ngủ chu kỳ gần nhất.
Tab <b>Power/Sleep</b> vẽ Thức vs Ngủ. <b>Quan trọng:</b> dù chọn deep sleep, khi node
<b>mất kết nối hoặc còn dữ liệu offline chưa gửi</b> thì node TỰ ĐỘNG dùng <b>light sleep</b>
để GIỮ buffer (RAM không mất); chỉ deep sleep khi đã online &amp; buffer rỗng. Khi offline,
node còn <b>backoff</b> (ngủ lâu dần tới 30s, retry ít) để tiết kiệm điện → vì vậy CPU offline
sẽ thấp dần. Chọn chế độ: <i>menuconfig → Node Sensor Configuration → Che do ngu</i>.</td></tr>
<tr><td><b style="color:#ce93d8">Watchdog</b></td>
<td>Lý do reset gần nhất · số lần boot (lưu trong RTC) · ⚠ số lần bị watchdog reset.
Bấm vào để xem chi tiết. <b>Test:</b> chập GPIO0 của node xuống GND (chạm nhẹ rồi thả)
để giả lập treo — watchdog sẽ reset node, reason đổi thành TASK_WDT, boot count tăng.</td></tr>
</table>

<h3 style="color:#ffd166">2b. Bộ lọc dữ liệu (raw vs filtered)</h3>
<p>Node <b>lọc dữ liệu cảm biến ngay tại chỗ</b> (edge processing), <b>mỗi tín hiệu một bộ lọc
phù hợp đặc tính vật lý</b> (chọn trong <i>NodeSensor_test/main/main.c</i>):</p>
<ul>
<li><b>Nhiệt độ &amp; Độ ẩm (SHT30) → EMA</b>: thay đổi chậm (quán tính nhiệt lớn), EMA bám
xu hướng &amp; mượt, chỉ tốn 1 biến RAM.</li>
<li><b>Áp suất (BMP180) → Moving Average</b> (cửa sổ nhỏ): ủi phẳng gợn nhiễu điện tử mà
không làm phản ứng chậm.</li>
<li><b>Ánh sáng (BH1750) → Median</b>: gạt nhiễu đột biến (bóng/côn trùng bay qua…), lấy
mức sáng nền thực tế.</li>
</ul>
<p>Nhãn <b>Filter:</b> hiển thị bản đồ bộ lọc (vd <i>T=EMA H=EMA P=MovingAvg L=Median</i>).</p>
<p>Hai tab <b>Climate</b> và <b>Environment</b> vẽ <b>cả hai</b>: đường <b>nét đứt = RAW</b>
(chưa lọc) và đường <b>nét liền đậm = FILTERED</b> (đã lọc) — để so sánh trực tiếp.
Ô <b>Plot:</b> chọn hiển thị <i>Both / Filtered only / Raw only</i>. File Excel xuất cả
cột raw (t,h,press,lux) lẫn filtered (tf,hf,pressf,luxf) để phân tích.</p>

<h3 style="color:#ffd166">3. Bộ đếm thông lượng (thanh phía trên trái)</h3>
<ul>
<li><b>sample/s</b>: số bản tin telemetry nhận được mỗi giây.</li>
<li><b>RX / Sent</b>: số gói gateway giải mã được / số gói node đã gửi (suy ra từ seq).</li>
<li><b>Delivery %</b>: RX / Sent của node đang chọn.</li>
<li><b>All nodes</b>: tổng RX/Sent của tất cả node.</li>
</ul>

<h3 style="color:#ffd166">4. Gateway Deep Profiler &amp; RTOS Profiler của node</h3>
<p>Panel Gateway: Heap (rò rỉ), %CPU core 0/1 của gateway, uptime, số node online, WiFi, Firebase.</p>
<p>Nút <b>🧩 RTOS</b> (góc trên) mở cửa sổ profiler của NODE: <b>%CPU từng core 0/1</b> và
<b>bảng các task FreeRTOS</b> đang chạy (tên, %CPU, core, stack còn trống, độ ưu tiên, trạng thái)
— để kiểm tra RTOS hoạt động thế nào, task nào ngốn CPU, stack có gần tràn không.</p>

<h3 style="color:#ffd166">5. Biểu đồ Distance↔Loss</h3>
<p>Tương quan giữa khoảng cách (trục X) và tỉ lệ mất gói % (trục Y); mỗi node
một màu. Dùng để đánh giá vùng phủ sóng.</p>

<h3 style="color:#ffd166">6. Các kịch bản kiểm thử (đều đánh giá trên NODE)</h3>
<table cellpadding="5">
<tr><td><b>Link &amp; ACK Stability</b></td><td>Node có nhận ACK từ gateway (online=1, RSSI hợp lệ)
→ chứng minh node gửi được qua LoRa và đường truyền/ACK hoạt động đúng.</td></tr>
<tr><td><b>Latency / Response Time</b></td><td>RTT node↔gateway nằm trong ngân sách (&lt;4s).</td></tr>
<tr><td><b>Encryption Active (AES/ASCON)</b></td><td>Node mã hóa gói (thuật toán hợp lệ +
thời gian mã hóa &gt; 0).</td></tr>
<tr><td><b>Bandwidth (pkt/h)</b></td><td>Thông lượng node tự tính (&gt; 0).</td></tr>
<tr><td><b>Sensor Sanity Check</b></td><td>Giá trị SHT30/BMP180/BH1750 nằm trong khoảng hợp lý.</td></tr>
<tr><td><b>Memory Leak Watch</b></td><td>Theo dõi heap node ~15s, cảnh báo nếu rò rỉ.</td></tr>
<tr><td><b>Offline Buffering &amp; Recovery</b></td><td>Gây mất kết nối (rút LoRa gateway) → kiểm tra
node lưu offline (buffer tăng) → khôi phục → buffer về 0.</td></tr>
<tr><td><b>OTA Update Test</b></td><td><b>Chỉ cần cắm Gateway</b> (Node có thể treo trên cao/ở xa). Nhập <i>OTA URL</i>
(link firmware .bin), tool gửi lệnh OTA qua LoRa → Node tự bật WiFi tải firmware → verify SHA-256 → reboot.
Tool xác nhận <b>THÀNH CÔNG khi Node quay lại LoRa với firmware version mới</b> (nhớ tăng <i>NODE_FW_VERSION</i>
cho bản mới). Không cần cắm Node vào laptop.</td></tr>
<tr><td><b>LoRa Replay Attack</b></td><td>Cần cắm <b>Gateway</b>. Tool bảo Gateway BẮT một gói hợp lệ, chờ &gt;5s
(quá cửa sổ chống replay), rồi PHÁT LẠI gói cũ. PASS nếu Gateway <b>từ chối</b> gói cũ (age &gt; window) —
chứng minh cơ chế chống tấn công phát lại theo nhãn thời gian hoạt động.</td></tr>
</table>
<p>Báo cáo có <i>Verification Accuracy</i> = tỉ lệ test đạt, kèm thống kê per-node.
Các test chạy được CHỈ CẦN cắm node (không cần gateway).</p>

<h3 style="color:#ffd166">7. Lưu trữ offline (store-and-forward)</h3>
<p>Khi node <b>mất kết nối</b> với gateway (không nhận ACK), node tự lưu các bản
đo vào <b>bộ nhớ cục bộ (ring buffer FIFO)</b> kèm <b>thời gian thực</b> (đồng bộ
từ gateway qua SNTP). Khi buffer đầy, bản ghi <b>cũ nhất bị bỏ</b> để nhận bản mới.
Khi có kết nối lại, node <b>gửi lại toàn bộ</b> (gói có cờ <i>backlog</i>, giữ đúng
timestamp gốc) rồi <b>giải phóng</b> đến khi buffer = 0.</p>
<p>Bảng <b>OFFLINE STORE-AND-FORWARD</b> hiển thị: số bản ghi đang giữ / dung lượng,
% còn trống, byte đang chiếm, tổng đã lưu / đã gửi lại / đã bỏ, thanh % đầy, và
<b>nhật ký sự kiện</b> (STORE khi mất kết nối, FLUSH khi gửi lại, BACKLOG kèm thời gian
và giá trị của bản ghi, DRAINED khi đã gửi & giải phóng hết). Tab <b>Offline Buffer</b>
vẽ số bản ghi theo thời gian. Kịch bản test <b>Offline Buffering &amp; Recovery</b>:
gây mất kết nối → kiểm tra buffer tăng → khôi phục → kiểm tra buffer về 0.</p>

<h3 style="color:#ffd166">8. Ghi &amp; phát lại phiên (Record / Replay) + Excel</h3>
<ul>
<li><b>⏺ Record</b>: bắt đầu ghi; bấm lại (⏹ Stop &amp; Save) để lưu phiên ra file
<b>.eplog</b>. File này chứa toàn bộ telemetry theo thời gian.</li>
<li><b>Kéo-thả file .eplog vào cửa sổ</b> (hoặc 📂 Open Log) để <b>phát lại</b>: hiện thanh
REPLAY với <b>Play/Pause</b>, thanh tua, tốc độ (0.5x…Max). Mọi biểu đồ/tile/sự kiện
dựng lại y như lúc chạy thật — bạn vẫn <b>kéo xem, phóng to/thu nhỏ biểu đồ, bấm tile,
đổi tab, đổi Source</b> như xem trực tiếp. Bấm ✕ Exit Replay để quay lại live.</li>
<li><b>Export Excel (.xlsx)</b>: xuất toàn bộ dữ liệu phiên — mỗi node 1 sheet (chuỗi
thời gian đầy đủ), sheet Gateway, Events (offline/watchdog), và Tests — để thống kê.</li>
</ul>

<h3 style="color:#ffd166">9. Mẹo dùng nút bấm</h3>
<ul>
<li>Bấm vào <b>ô tile bất kỳ</b> để mở popup chi tiết (biểu đồ + min/max/avg).</li>
<li><b>Export HTML/XML</b>: xuất báo cáo đa node (kèm testsuite kiểu JUnit).</li>
<li><b>Clear Session</b>: xóa toàn bộ dữ liệu phiên hiện tại.</li>
</ul>
"""


class HelpDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("EdgeProfiler — Docs / Help")
        self.resize(820, 640)
        lay = QtWidgets.QVBoxLayout(self)
        view = QtWidgets.QTextBrowser()
        view.setOpenExternalLinks(True)
        view.setHtml(HELP_HTML)
        view.setStyleSheet("background:#0f1419; color:#e6e6e6; font-size:13px;")
        lay.addWidget(view)
        btn = QtWidgets.QPushButton("Đóng")
        btn.clicked.connect(self.accept)
        lay.addWidget(btn, 0, Qt.AlignRight)


# ==========================================================================
# FreeRTOS profiler dialog (per-task CPU + per-core)
# ==========================================================================
class RtosDialog(QtWidgets.QDialog):
    def __init__(self, get_node, parent=None):
        super().__init__(parent)
        self.get_node = get_node
        self.setWindowTitle("FreeRTOS Profiler — node tasks & cores")
        self.resize(680, 520)
        lay = QtWidgets.QVBoxLayout(self)

        self.lbl = QtWidgets.QLabel("Node —")
        self.lbl.setStyleSheet("font-weight:700; color:#4fc3f7;")
        lay.addWidget(self.lbl)

        cores = QtWidgets.QGridLayout()
        cores.addWidget(QtWidgets.QLabel("Core 0 (PRO)"), 0, 0)
        self.bar0 = QtWidgets.QProgressBar(); self.bar0.setRange(0, 100); self.bar0.setFormat("%p%")
        cores.addWidget(self.bar0, 0, 1)
        cores.addWidget(QtWidgets.QLabel("Core 1 (APP)"), 1, 0)
        self.bar1 = QtWidgets.QProgressBar(); self.bar1.setRange(0, 100); self.bar1.setFormat("%p%")
        cores.addWidget(self.bar1, 1, 1)
        lay.addLayout(cores)

        self.table = QtWidgets.QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(
            ["Task", "CPU %", "Core", "Stack free (B)", "Prio", "State"])
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        lay.addWidget(self.table, 1)

        note = QtWidgets.QLabel("CPU% = phần thời gian chạy của task (tích luỹ). "
                                "State: R=running r=ready B=blocked S=suspended. "
                                "Stack free = high-water mark (càng nhỏ càng sát tràn stack).")
        note.setWordWrap(True); note.setStyleSheet("color:#8aa; font-size:11px;")
        lay.addWidget(note)

        self.timer = QTimer(self); self.timer.timeout.connect(self.refresh)
        self.timer.start(500)
        self.refresh()

    def refresh(self):
        nd = self.get_node()
        if not nd:
            self.lbl.setText("No node-direct data (cắm cổng NODE để xem)")
            return
        self.lbl.setText(f"Node {nd.node_id}  ·  {len(nd.tasks)} tasks  ·  "
                         f"core0 {nd.core0}%  core1 {nd.core1}%")
        self.bar0.setValue(int(nd.core0)); self.bar1.setValue(int(nd.core1))
        tasks = sorted(nd.tasks, key=lambda t: -int(t.get("cpu", 0)))
        self.table.setRowCount(len(tasks))
        for r, t in enumerate(tasks):
            core = t.get("core", -1)
            vals = [str(t.get("n", "?")), f"{t.get('cpu', 0)}", "—" if core < 0 else str(core),
                    f"{t.get('stk', 0)}", f"{t.get('prio', 0)}", str(t.get("st", "?"))]
            for c, v in enumerate(vals):
                it = QtWidgets.QTableWidgetItem(v)
                if c == 1:
                    cpu = int(t.get("cpu", 0))
                    col = "#ff5c5c" if cpu > 60 else "#ffd166" if cpu > 25 else "#26c281"
                    it.setForeground(QtGui.QColor(col))
                if c == 3 and int(t.get("stk", 9999)) < 400:
                    it.setForeground(QtGui.QColor("#ff5c5c"))   # stack gần tràn
                self.table.setItem(r, c, it)

    def closeEvent(self, ev):
        self.timer.stop(); ev.accept()


# ==========================================================================
# Test sequencer
# ==========================================================================
class TestSequencer(QThread):
    log = pyqtSignal(str)
    result = pyqtSignal(dict)
    done = pyqtSignal()

    def __init__(self, get_gw_worker, get_state, scenario, extra=None, parent=None):
        super().__init__(parent)
        self.get_gw = get_gw_worker          # callable -> SerialWorker or None
        self.get_state = get_state           # callable -> dict snapshot
        self.scenario = scenario
        self.extra = extra or {}             # tham so phu (vd URL OTA)
        self._abort = False

    def abort(self):
        self._abort = True

    def _ts(self):
        return datetime.datetime.now().strftime("%H:%M:%S")

    def _send(self, obj):
        w = self.get_gw()
        if not w:
            self.log.emit(f"[{self._ts()}] ! No gateway port connected")
            return False
        self.log.emit(f"[{self._ts()}] TX -> {json.dumps(obj)}")
        return w.send_command(obj)

    def _wait(self, pred, timeout, desc):
        self.log.emit(f"[{self._ts()}] WAIT: {desc} (<= {timeout}s)")
        end = time.time() + timeout
        while time.time() < end:
            if self._abort:
                return False, {}
            st = self.get_state()
            if pred(st):
                return True, st
            self.msleep(120)
        return False, self.get_state()

    def _res(self, name, ok, detail):
        self.log.emit(f"[{self._ts()}] {'PASS' if ok else 'FAIL'}: {name} -- {detail}")
        self.result.emit({"name": name, "passed": bool(ok), "detail": detail,
                          "ts": datetime.datetime.now().isoformat(timespec="seconds")})

    # --- node-centric scenarios (evaluate the NODE; no gateway link needed) ---
    def _t_link(self):
        n = "Link & ACK Stability"
        self.log.emit(f"\n=== {n} ===")
        self.log.emit(f"[{self._ts()}] Checking node gets ACKs from gateway (LoRa link OK)…")
        ok, st = self._wait(lambda s: s.get("online") == 1 and (s.get("rssi") is not None),
                            10, "node online + valid ACK/RSSI")
        deliv = (100.0 * st.get("rx", 0) / st.get("sent", 1)) if st.get("sent") else 0
        self._res(n, ok, f"online={st.get('online')} rssi={st.get('rssi')} "
                         f"delivery={deliv:.0f}% (proves LoRa TX + gateway ACK works)")
        return ok

    def _t_sensor(self):
        n = "Sensor Sanity Check"
        self.log.emit(f"\n=== {n} ===")
        def ok_ranges(s):
            t, h, p = s.get("t"), s.get("h"), s.get("press")
            return (isinstance(t, (int, float)) and -40 <= t <= 85 and
                    isinstance(h, (int, float)) and 0 <= h <= 100 and
                    isinstance(p, (int, float)) and 800 <= p <= 1100)
        ok, st = self._wait(ok_ranges, 8, "SHT30/BMP180/BH1750 values within plausible ranges")
        self._res(n, ok, f"T={st.get('t')} H={st.get('h')} P={st.get('press')}")
        return ok

    def _t_memory(self):
        n = "Memory Leak Watch"
        self.log.emit(f"\n=== {n} ===")
        self.log.emit(f"[{self._ts()}] Sampling node heap for ~15s…")
        samples = []
        end = time.time() + 15
        while time.time() < end and not self._abort:
            h = self.get_state().get("heap")
            if isinstance(h, (int, float)) and h > 0:
                samples.append(h)
            self.msleep(700)
        if len(samples) < 3:
            self._res(n, False, "not enough heap samples"); return False
        drift = samples[-1] - samples[0]
        ok = drift >= -2048          # allow tiny jitter; flag steady downward leak
        self._res(n, ok, f"heap start={samples[0]:.0f} end={samples[-1]:.0f} "
                         f"drift={drift:+.0f} B {'(stable)' if ok else '(possible leak!)'}")
        return ok

    def _t_power(self):
        n = "Power / Sleep Mode"
        self.log.emit(f"\n=== {n} ===")
        ok, st = self._wait(
            lambda s: isinstance(s.get("sleep_ms"), (int, float)) and s.get("sleep_ms") > 0
                      and s.get("sleep_mode") in ("light", "deep"),
            10, "node sleeps between cycles (light/deep)")
        self._res(n, ok, f"mode={st.get('sleep_mode')} sleep={st.get('sleep_ms')}ms "
                         f"awake={st.get('awake_ms')}ms CPU={st.get('cpu')}%")
        return ok

    def _t_wdt(self):
        n = "Watchdog Recovery"
        self.log.emit(f"\n=== {n} ===")
        st0 = self.get_state()
        if not st0.get("wdt_s"):
            self._res(n, False, "node not reporting watchdog status"); return False
        boot0, wr0 = st0.get("boot", 0), st0.get("wdt_resets", 0)
        self.log.emit(f"[{self._ts()}] Watchdog timeout = {st0.get('wdt_s')}s. "
                      f">>> CHẬP GPIO0 của node xuống GND ngay để giả lập TREO…")
        ok, st = self._wait(
            lambda s: s.get("wdt_resets", 0) > wr0 or
                      (s.get("boot", 0) > boot0 and s.get("reset") in ("TASK_WDT", "INT_WDT", "WDT")),
            45, "node hangs, watchdog fires, node reboots & recovers")
        if not ok:
            self._res(n, False, "no watchdog reset observed (short GPIO0→GND during the test)")
            return False
        # confirm it comes back online
        ok2, st2 = self._wait(lambda s: s.get("online") == 1, 20, "node back online after reset")
        self._res(n, ok2, f"reset={st.get('reset')} boot {boot0}->{st.get('boot')} "
                          f"wdt_resets={st.get('wdt_resets')} recovered={ok2}")
        return ok2

    def _t_offline(self):
        n = "Offline Buffering & Recovery"
        self.log.emit(f"\n=== {n} ===")
        self.log.emit(f"[{self._ts()}] >>> INDUCE AN OUTAGE NOW: power off the gateway OR move/shield "
                      f"the node so it loses the link for a few seconds, then restore it.")
        ok1, st = self._wait(lambda s: s.get("buf_count", 0) > 0, 45,
                             "node stores data offline (buffer grows)")
        if not ok1:
            self._res(n, False, "no offline buffering detected (induce a link outage during the test)")
            return False
        self.log.emit(f"[{self._ts()}] buffered {st.get('buf_count')} record(s); now RESTORE the link…")
        ok2, st2 = self._wait(lambda s: s.get("buf_count", 99) == 0, 120,
                              "buffer drains to 0 and memory is freed after reconnect")
        self._res(n, ok2, f"final buffer={st2.get('buf_count')} "
                          f"flushed={st2.get('buf_flushed')} dropped={st2.get('buf_dropped')}")
        return ok2

    def _t_bandwidth(self):
        n = "Bandwidth (pkt/h)"
        self.log.emit(f"\n=== {n} ===")
        ok, st = self._wait(lambda s: isinstance(s.get("pph"), (int, float)) and s.get("pph") > 0,
                            8, "node reports packets/hour")
        pph = st.get("pph") or 0
        self._res(n, ok, f"throughput ~{pph:.0f} pkt/h (node-computed)")
        return ok

    def _t_crypto(self):
        n = "Encryption Active (AES/ASCON)"
        self.log.emit(f"\n=== {n} ===")
        ok, st = self._wait(
            lambda s: s.get("algo", "NONE") in ("AES-128-GCM", "ASCON-128")
                      and isinstance(s.get("enc"), (int, float)) and s.get("enc") > 0,
            6, "packets encrypted (valid algo + non-zero encrypt time)")
        self._res(n, ok, f"algo={st.get('algo')} enc_time={st.get('enc')} µs")
        return ok

    def _t_latency(self):
        n = "Latency / Response Time"
        self.log.emit(f"\n=== {n} ===")
        ok, st = self._wait(
            lambda s: isinstance(s.get("rtt"), (int, float)) and 0 < s.get("rtt") < 4000,
            8, "node↔gateway RTT within budget (<4s)")
        self._res(n, ok, f"rtt={st.get('rtt')} ms")
        return ok

    # --- gateway-side scenarios (can cam cong GATEWAY) ---
    def _t_ota(self):
        # OTA tu xa: CHI CAN CAM GATEWAY. Node co the treo tren cao, cach xa gateway.
        # Xac nhan thanh cong bang cach cho node quay lai LoRa voi fw_ver LON HON.
        n = "OTA Update Test"
        self.log.emit(f"\n=== {n} (tu xa - chi can cam GATEWAY) ===")
        url = self.extra.get("ota_url", "").strip()
        if not self.get_gw():
            self._res(n, False, "Chua cam cong GATEWAY"); return False
        if not url:
            self._res(n, False, "Chua nhap 'OTA URL' (o tren nut RUN)"); return False
        st0 = self.get_state()
        node = st0.get("gw_node_id") or st0.get("node_id") or 1
        from_ver = int(st0.get("gw_fw_ver") or 0)
        if from_ver == 0:
            self._res(n, False, "Chua thay node qua Gateway - node phai ONLINE (dang gui LoRa) truoc khi OTA")
            return False
        self.log.emit(f"[{self._ts()}] Node {node} dang chay firmware v{from_ver}.")
        self.log.emit(f"[{self._ts()}] LUU Y: firmware moi PHAI co NODE_FW_VERSION > {from_ver} de tool xac nhan duoc.")
        self._send({"cmd": "send_cmd", "node": int(node), "op": 3, "url": url})
        self.log.emit(f"[{self._ts()}] Da gui lenh OTA -> node {node}: {url}")
        self.log.emit(f"[{self._ts()}] Node se OFFLINE ~30-90s (bat WiFi tai firmware) roi quay lai. Dang cho...")
        # Chi can GATEWAY: node quay lai voi fw_ver moi => OTA thanh cong
        ok, st = self._wait(lambda s: int(s.get("gw_fw_ver") or 0) > from_ver, 240,
                            f"node OTA xong & quay lai LoRa voi firmware v>{from_ver}")
        new_ver = int(st.get("gw_fw_ver") or from_ver)
        if ok:
            self._res(n, True, f"OTA THANH CONG: node {node} len firmware v{from_ver} -> v{new_ver} "
                               f"(xac nhan qua Gateway, KHONG can cam node)")
        else:
            self._res(n, False, f"Sau 240s van thay v{from_ver}: OTA that bai/rollback, node chua nhan lenh, "
                                f"chua tang NODE_FW_VERSION, hoac node khong vao duoc WiFi")
        return ok

    def _t_replay(self):
        n = "LoRa Replay Attack"
        self.log.emit(f"\n=== {n} ===")
        if not self.get_gw():
            self._res(n, False, "Chua cam cong GATEWAY (replay test chay tren gateway)"); return False
        # 1) bat 1 goi hop le tu node
        self._send({"cmd": "replay_capture"})
        self.log.emit(f"[{self._ts()}] Yeu cau gateway BAT 1 goi hop le tu node (cho node uplink)...")
        self.msleep(1500)
        # 2) doi qua cua so chong replay (~6s) roi PHAT LAI goi cu
        self.log.emit(f"[{self._ts()}] Cho 6s (qua cua so chong replay) roi phat lai goi cu...")
        for _ in range(60):
            if self._abort:
                return False
            self.msleep(100)
        before = self.get_state().get("replay")
        self._send({"cmd": "replay_now"})
        ok, st = self._wait(
            lambda s: (s.get("replay") is not before) and (s.get("replay") or {}).get("result"),
            8, "gateway phan hoi ket qua replay")
        rr = st.get("replay") or {}
        blocked = rr.get("result") == "blocked"
        self._res(n, blocked, f"age={rr.get('age_s')}s window={rr.get('window_s')}s -> "
                  f"{str(rr.get('result','?')).upper()} " +
                  ("PASS: goi phat lai (cu) BI TU CHOI — chong replay hoat dong"
                   if blocked else "goi con trong cua so, thu lai (doi lau hon)"))
        return blocked

    def run(self):
        self.log.emit(f"\n######## START: {self.scenario} ########")
        s = self.scenario
        gw_scn = s.startswith("OTA") or s.startswith("LoRa Replay")
        if gw_scn:
            if not self.get_state().get("has_gw"):
                self._res(self.scenario, False, "Chua cam cong GATEWAY cho kich ban nay")
                self.done.emit(); return
        elif not self.get_state().get("has_node"):
            self._res(self.scenario, False, "No node telemetry — connect the NODE's USB port")
            self.done.emit(); return
        try:
            if s.startswith("Link"):
                self._t_link()
            elif s.startswith("Latency"):
                self._t_latency()
            elif s.startswith("Encryption"):
                self._t_crypto()
            elif s.startswith("Bandwidth"):
                self._t_bandwidth()
            elif s.startswith("Sensor"):
                self._t_sensor()
            elif s.startswith("Memory"):
                self._t_memory()
            elif s.startswith("Power"):
                self._t_power()
            elif s.startswith("Watchdog"):
                self._t_wdt()
            elif s.startswith("Offline Buffering"):
                self._t_offline()
            elif s.startswith("OTA"):
                self._t_ota()
            elif s.startswith("LoRa Replay"):
                self._t_replay()
            elif s.startswith("Auto-Run"):
                for f in (self._t_link, self._t_latency, self._t_crypto,
                          self._t_bandwidth, self._t_sensor, self._t_memory, self._t_power):
                    if self._abort:
                        break
                    f()
        except Exception as exc:                           # noqa: BLE001
            self._res(self.scenario, False, f"exception: {exc}")
        self.log.emit("######## SEQUENCE COMPLETE ########\n")
        self.done.emit()


# ==========================================================================
# Main window
# ==========================================================================
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{APP_NAME} v{APP_VERSION} - ESP32-S3 LoRa HIL QA")
        self.resize(1640, 980)

        self.workers = {}                 # port -> SerialWorker
        self.nodes = {}                   # node_id -> NodeData (DIRECT from node USB)
        self.nodes_gw = {}                # node_id -> NodeData (RELAYED by gateway)
        self.gw = GatewayData()
        self.gw_port = None               # port that behaves as the gateway
        self.source_mode = "Node (direct)"
        self.plot_mode = "Both (raw+filtered)"
        self.frozen = False
        self.seq = None
        self.detail_dialogs = []
        self.session_start = datetime.datetime.now()
        self.test_results = []
        self.anomaly_injected = 0
        self.anomaly_detected = 0
        # OTA + security (replay) tracking
        self.ota_state = {}          # node_id -> latest OTA progress dict
        self.sec_events = deque(maxlen=200)
        self.replay_result = None    # latest {"event":"replay_test",...}

        # session recording / replay
        self.recording = False
        self.record_frames = []          # [(t_rel, kind, port, data)]
        self.record_t0 = None
        self.replay_active = False
        self.replay_paused = True
        self.replay_frames = []
        self.replay_pos = 0
        self.replay_time = 0.0
        self.replay_dur = 0.0
        self.replay_speed = 1.0

        self.setAcceptDrops(True)         # drag a .eplog file onto the window
        self._build_ui()

        self.timer = QTimer(self)
        self.timer.timeout.connect(self._refresh_views)
        self.timer.start(300)
        self.play_timer = QTimer(self)
        self.play_timer.timeout.connect(self._replay_tick)
        self.play_timer.start(100)
        self._refresh_ports()

    # ------------------------------------------------------------------ UI
    def _build_ui(self):
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        root = QtWidgets.QVBoxLayout(central)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(8)
        root.addWidget(self._build_connbar())

        split = QtWidgets.QSplitter(Qt.Horizontal)
        split.addWidget(self._build_left())
        split.addWidget(self._build_right())
        split.setStretchFactor(0, 3)
        split.setStretchFactor(1, 2)
        root.addWidget(split, 1)
        root.addWidget(self._build_replaybar())
        root.addWidget(self._build_bottombar())

    def _build_connbar(self):
        box = QtWidgets.QFrame(); box.setObjectName("Card")
        lay = QtWidgets.QHBoxLayout(box); lay.setContentsMargins(10, 6, 10, 6)
        lay.addWidget(QtWidgets.QLabel("Port:"))
        self.cmb_port = QtWidgets.QComboBox(); self.cmb_port.setMinimumWidth(240)
        lay.addWidget(self.cmb_port)
        b = QtWidgets.QPushButton("↻"); b.setFixedWidth(32); b.clicked.connect(self._refresh_ports)
        lay.addWidget(b)
        lay.addWidget(QtWidgets.QLabel("Baud:"))
        self.cmb_baud = QtWidgets.QComboBox()
        self.cmb_baud.addItems(["115200", "230400", "921600", "57600", "9600"])
        lay.addWidget(self.cmb_baud)
        self.btn_add = QtWidgets.QPushButton("+ Connect Port"); self.btn_add.setObjectName("Primary")
        self.btn_add.clicked.connect(self._add_port)
        lay.addWidget(self.btn_add)

        lay.addWidget(QtWidgets.QLabel("  Connected:"))
        self.lbl_ports = QtWidgets.QLabel("none")
        self.lbl_ports.setStyleSheet("color:#26c281;")
        lay.addWidget(self.lbl_ports)
        lay.addStretch(1)

        self.btn_rec = QtWidgets.QPushButton("⏺ Record")
        self.btn_rec.setCheckable(True)
        self.btn_rec.toggled.connect(self._toggle_record)
        self.btn_rec.setToolTip("Start recording; press again to stop & save a .eplog session file")
        lay.addWidget(self.btn_rec)
        self.btn_openlog = QtWidgets.QPushButton("📂 Open Log")
        self.btn_openlog.setToolTip("Open a saved .eplog to replay (or drag the file onto the window)")
        self.btn_openlog.clicked.connect(self._open_log)
        lay.addWidget(self.btn_openlog)

        self.btn_rtos = QtWidgets.QPushButton("🧩 RTOS")
        self.btn_rtos.setToolTip("Xem task FreeRTOS + %CPU từng core của node")
        self.btn_rtos.clicked.connect(self._open_rtos)
        lay.addWidget(self.btn_rtos)
        self.btn_help = QtWidgets.QPushButton("❔ Docs / Help")
        self.btn_help.clicked.connect(self._open_help)
        lay.addWidget(self.btn_help)
        self.btn_freeze = QtWidgets.QPushButton("⏸ Freeze")
        self.btn_freeze.setCheckable(True)
        self.btn_freeze.toggled.connect(self._toggle_freeze)
        lay.addWidget(self.btn_freeze)
        self.btn_discon = QtWidgets.QPushButton("Disconnect All")
        self.btn_discon.clicked.connect(self._disconnect_all)
        lay.addWidget(self.btn_discon)
        return box

    def _build_left(self):
        w = QtWidgets.QWidget()
        lay = QtWidgets.QVBoxLayout(w); lay.setContentsMargins(0, 0, 0, 0)

        # node selector row
        row = QtWidgets.QHBoxLayout()
        row.addWidget(QtWidgets.QLabel("Source:"))
        self.cmb_src = QtWidgets.QComboBox(); self.cmb_src.setMinimumWidth(140)
        self.cmb_src.addItems(["Node (direct)", "Gateway (relayed)", "Both (compare)"])
        self.cmb_src.currentTextChanged.connect(self._on_source_change)
        row.addWidget(self.cmb_src)
        row.addWidget(QtWidgets.QLabel("View node:"))
        self.cmb_node = QtWidgets.QComboBox(); self.cmb_node.setMinimumWidth(110)
        self.cmb_node.currentIndexChanged.connect(lambda *_: self._refresh_views())
        row.addWidget(self.cmb_node)
        row.addWidget(QtWidgets.QLabel("Algo:"))
        self.lbl_algo = QtWidgets.QLabel("-"); self.lbl_algo.setStyleSheet("color:#80d8ff;font-weight:700;")
        row.addWidget(self.lbl_algo)
        row.addWidget(QtWidgets.QLabel("  Filter:"))
        self.lbl_filter = QtWidgets.QLabel("-"); self.lbl_filter.setStyleSheet("color:#a0e060;font-weight:700;")
        row.addWidget(self.lbl_filter)
        row.addWidget(QtWidgets.QLabel("Plot:"))
        self.cmb_plotmode = QtWidgets.QComboBox()
        self.cmb_plotmode.addItems(["Both (raw+filtered)", "Filtered only", "Raw only"])
        self.cmb_plotmode.currentTextChanged.connect(lambda t: setattr(self, "plot_mode", t))
        row.addWidget(self.cmb_plotmode)
        row.addStretch(1)
        self.lbl_fsm = QtWidgets.QLabel("UNKNOWN"); self.lbl_fsm.setObjectName("FsmSafe")
        self.lbl_fsm.setAlignment(Qt.AlignCenter); self.lbl_fsm.setMinimumWidth(130)
        row.addWidget(self.lbl_fsm)
        lay.addLayout(row)

        # throughput / counters + distance calibration
        stat = QtWidgets.QFrame(); stat.setObjectName("Card")
        sl = QtWidgets.QHBoxLayout(stat); sl.setContentsMargins(10, 5, 10, 5)
        self.lbl_rate = QtWidgets.QLabel("– sample/s")
        self.lbl_rx = QtWidgets.QLabel("RX – / Sent –")
        self.lbl_deliv = QtWidgets.QLabel("Delivery –")
        self.lbl_tot = QtWidgets.QLabel("All nodes: RX –")
        self.lbl_compare = QtWidgets.QLabel("")
        for w_, col in ((self.lbl_rate, "#80d8ff"), (self.lbl_rx, "#a0e060"),
                        (self.lbl_deliv, "#26c281"), (self.lbl_tot, "#ffd166"),
                        (self.lbl_compare, "#ff8a65")):
            w_.setStyleSheet(f"color:{col}; font-weight:700;")
            sl.addWidget(w_)
        sl.addStretch(1)
        sl.addWidget(QtWidgets.QLabel("Dist calib  RSSI@1m:"))
        self.spin_ref = QtWidgets.QDoubleSpinBox(); self.spin_ref.setRange(-120, 0)
        self.spin_ref.setValue(CALIB["ref"]); self.spin_ref.setSuffix(" dBm")
        self.spin_ref.valueChanged.connect(lambda v: CALIB.update(ref=v))
        sl.addWidget(self.spin_ref)
        sl.addWidget(QtWidgets.QLabel("n:"))
        self.spin_n = QtWidgets.QDoubleSpinBox(); self.spin_n.setRange(1.5, 5.0)
        self.spin_n.setSingleStep(0.1); self.spin_n.setValue(CALIB["n"])
        self.spin_n.valueChanged.connect(lambda v: CALIB.update(n=v))
        sl.addWidget(self.spin_n)
        self.spin_cal_d = QtWidgets.QDoubleSpinBox(); self.spin_cal_d.setRange(0.1, 5000)
        self.spin_cal_d.setValue(1.0); self.spin_cal_d.setSuffix(" m")
        sl.addWidget(self.spin_cal_d)
        btn_cal = QtWidgets.QPushButton("Calibrate@here")
        btn_cal.setToolTip("Set RSSI@1m from the selected node's current RSSI and the distance you typed")
        btn_cal.clicked.connect(self._calibrate_distance)
        sl.addWidget(btn_cal)
        lay.addWidget(stat)

        # metric tiles
        tiles = QtWidgets.QGridLayout(); tiles.setSpacing(6)
        self.tiles = {}
        specs = [("rtt", "Response (RTT)", "ms", C["rtt"]),
                 ("dist", "Distance", "m", C["dist"]),
                 ("pdr", "PDR", "%", C["pdr"]),
                 ("pph", "Bandwidth", "pkt/h", C["pph"]),
                 ("proc", "Proc time", "µs", C["proc"]),
                 ("enc", "Encrypt time", "µs", C["enc"]),
                 ("rssi", "RSSI", "dBm", C["rssi"]),
                 ("heap", "Node heap", "B", C["heap"]),
                 ("cpu", "Node CPU", "%", "#ff7043"),
                 ("sleep", "Sleep", "mode · ms", "#80cbc4"),
                 ("wdt", "Watchdog", "reset · boot", "#ce93d8")]
        for i, (k, ti, u, col) in enumerate(specs):
            t = MetricTile(k, ti, u, col)
            t.clicked.connect(self._open_detail)
            self.tiles[k] = t
            tiles.addWidget(t, i // 4, i % 4)
        lay.addLayout(tiles)

        # plot tabs
        self.tabs = QtWidgets.QTabWidget()
        self.p_climate = self._dual("Climate  (raw = dashed · filtered = solid)",
                                    "Temp (°C)", C["t"], "Humidity (%)", C["h"])
        self.p_env = self._dual("Environment  (raw = dashed · filtered = solid)",
                                "Pressure (hPa)", C["press"], "Light (lux)", C["lux"])
        self.p_net = self._dual("LoRa Link", "RSSI (dBm)", C["rssi"], "SNR (dB)", C["snr"])
        self.p_perf = self._dual("Performance", "RTT (ms)", C["rtt"], "Distance (m)", C["dist"])
        self.p_corr = self._scatter()
        self.p_buf = self._single("Offline Buffer occupancy", "records buffered", "#ff8a65")
        self.p_power = self._dual("Power / Sleep per cycle", "Awake (ms)", "#ff7043",
                                  "Sleep (ms)", "#80cbc4")
        self.tabs.addTab(self.p_climate, "Climate")
        self.tabs.addTab(self.p_env, "Environment")
        self.tabs.addTab(self.p_net, "Network")
        self.tabs.addTab(self.p_perf, "Performance")
        self.tabs.addTab(self.p_corr, "Distance↔Loss")
        self.tabs.addTab(self.p_buf, "Offline Buffer")
        self.tabs.addTab(self.p_power, "Power/Sleep")
        lay.addWidget(self.tabs, 1)
        return w

    def _build_right(self):
        w = QtWidgets.QWidget()
        lay = QtWidgets.QVBoxLayout(w); lay.setContentsMargins(0, 0, 0, 0)

        # gateway profiler
        gwbox = QtWidgets.QFrame(); gwbox.setObjectName("Card")
        gl = QtWidgets.QVBoxLayout(gwbox)
        h = QtWidgets.QLabel("GATEWAY DEEP PROFILER"); h.setObjectName("Header")
        gl.addWidget(h)
        self.p_heap = pg.PlotWidget(); self.p_heap.setBackground(BG_NORMAL)
        self.p_heap.setTitle("Gateway Free Heap (leak watch)", color="#ddd", size="9pt")
        self.p_heap.showGrid(x=True, y=True, alpha=0.2)
        self.c_heap = self.p_heap.plot(pen=pg.mkPen(C["heap"], width=2))
        self.p_heap.setMaximumHeight(170)
        gl.addWidget(self.p_heap)
        grid = QtWidgets.QGridLayout()
        self.bar_c0 = self._bar(); self.bar_c1 = self._bar()
        grid.addWidget(QtWidgets.QLabel("Core 0"), 0, 0); grid.addWidget(self.bar_c0, 0, 1)
        grid.addWidget(QtWidgets.QLabel("Core 1"), 1, 0); grid.addWidget(self.bar_c1, 1, 1)
        self.lbl_gw = QtWidgets.QLabel("heap - | up - | nodes - | wifi - | firebase -")
        self.lbl_gw.setStyleSheet("color:#9fb;")
        grid.addWidget(self.lbl_gw, 2, 0, 1, 2)
        gl.addLayout(grid)
        lay.addWidget(gwbox)

        # nodes overview table
        nbox = QtWidgets.QFrame(); nbox.setObjectName("Card")
        nl = QtWidgets.QVBoxLayout(nbox)
        h2 = QtWidgets.QLabel("NODES OVERVIEW"); h2.setObjectName("Header"); nl.addWidget(h2)
        self.table = QtWidgets.QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels(
            ["Node", "T°C", "H%", "RSSI", "PDR%", "RTT", "Dist m", "FSM", "Online"])
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setMaximumHeight(160)
        nl.addWidget(self.table)
        lay.addWidget(nbox)

        # offline storage monitor
        obox = QtWidgets.QFrame(); obox.setObjectName("Card")
        ol = QtWidgets.QVBoxLayout(obox)
        oh = QtWidgets.QLabel("OFFLINE STORE-AND-FORWARD"); oh.setObjectName("Header")
        ol.addWidget(oh)
        og = QtWidgets.QGridLayout()
        self.lbl_buf_use = QtWidgets.QLabel("0 / 0 records")
        self.lbl_buf_free = QtWidgets.QLabel("100% free")
        self.lbl_buf_bytes = QtWidgets.QLabel("0 B used")
        self.lbl_buf_tot = QtWidgets.QLabel("stored 0 · flushed 0 · dropped 0")
        for w_, c in ((self.lbl_buf_use, "#ff8a65"), (self.lbl_buf_free, "#26c281"),
                      (self.lbl_buf_bytes, "#ffd166"), (self.lbl_buf_tot, "#8aa")):
            w_.setStyleSheet(f"color:{c}; font-weight:700;")
        self.buf_bar = QtWidgets.QProgressBar(); self.buf_bar.setRange(0, 100); self.buf_bar.setValue(0)
        self.buf_bar.setFormat("buffer %p% full")
        og.addWidget(self.lbl_buf_use, 0, 0); og.addWidget(self.lbl_buf_free, 0, 1)
        og.addWidget(self.lbl_buf_bytes, 1, 0); og.addWidget(self.lbl_buf_tot, 1, 1)
        og.addWidget(self.buf_bar, 2, 0, 1, 2)
        ol.addLayout(og)
        ol.addWidget(QtWidgets.QLabel("Event log (store / flush / backlog):"))
        self.off_log = QtWidgets.QPlainTextEdit(); self.off_log.setReadOnly(True)
        self.off_log.setObjectName("Console"); self.off_log.setMaximumHeight(120)
        ol.addWidget(self.off_log)
        lay.addWidget(obox)

        # test sequencer
        tbox = QtWidgets.QFrame(); tbox.setObjectName("Card")
        tl = QtWidgets.QVBoxLayout(tbox)
        h3 = QtWidgets.QLabel("AUTOMATED TEST SEQUENCER"); h3.setObjectName("Header"); tl.addWidget(h3)
        rr = QtWidgets.QHBoxLayout()
        self.cmb_scn = QtWidgets.QComboBox()
        self.cmb_scn.addItems([
            "Link & ACK Stability",
            "Latency / Response Time",
            "Encryption Active (AES/ASCON)",
            "Bandwidth (pkt/h)",
            "Sensor Sanity Check",
            "Memory Leak Watch",
            "Power / Sleep Mode",
            "Watchdog Recovery",
            "Offline Buffering & Recovery",
            "OTA Update Test",
            "LoRa Replay Attack",
            "Auto-Run Node Suite",
        ])
        rr.addWidget(self.cmb_scn, 1)
        self.btn_run = QtWidgets.QPushButton("RUN"); self.btn_run.setObjectName("Primary")
        self.btn_run.clicked.connect(self._run_test); rr.addWidget(self.btn_run)
        tl.addLayout(rr)
        # Hang tham so cho OTA (URL firmware .bin) — dung cho kich ban "OTA Update Test"
        rr2 = QtWidgets.QHBoxLayout()
        rr2.addWidget(QtWidgets.QLabel("OTA URL:"))
        self.ed_ota_url = QtWidgets.QLineEdit()
        self.ed_ota_url.setPlaceholderText("http://<server>/node_fw.bin  (cho kich ban OTA Update Test)")
        rr2.addWidget(self.ed_ota_url, 1)
        tl.addLayout(rr2)
        self.console = QtWidgets.QPlainTextEdit(); self.console.setReadOnly(True)
        self.console.setObjectName("Console"); self.console.setMaximumBlockCount(6000)
        tl.addWidget(self.console, 1)
        lay.addWidget(tbox, 1)
        return w

    def _build_replaybar(self):
        box = QtWidgets.QFrame(); box.setObjectName("Card")
        box.setVisible(False)
        self.replay_bar = box
        lay = QtWidgets.QHBoxLayout(box); lay.setContentsMargins(10, 5, 10, 5)
        tag = QtWidgets.QLabel("▶ REPLAY")
        tag.setStyleSheet("color:#ffd166; font-weight:700;")
        lay.addWidget(tag)
        self.btn_play = QtWidgets.QPushButton("▶ Play"); self.btn_play.setObjectName("Primary")
        self.btn_play.clicked.connect(self._replay_play_pause)
        lay.addWidget(self.btn_play)
        self.slider = QtWidgets.QSlider(Qt.Horizontal)
        self.slider.setRange(0, 1000)
        self.slider.valueChanged.connect(self._replay_seek)
        lay.addWidget(self.slider, 1)
        self.lbl_time = QtWidgets.QLabel("00:00 / 00:00")
        self.lbl_time.setStyleSheet("font-family:Consolas,monospace;")
        lay.addWidget(self.lbl_time)
        lay.addWidget(QtWidgets.QLabel("Speed:"))
        self.cmb_speed = QtWidgets.QComboBox()
        self.cmb_speed.addItems(["0.5x", "1x", "2x", "4x", "10x", "Max"])
        self.cmb_speed.setCurrentText("1x")
        self.cmb_speed.currentTextChanged.connect(self._replay_speed_change)
        lay.addWidget(self.cmb_speed)
        btn_exit = QtWidgets.QPushButton("✕ Exit Replay")
        btn_exit.clicked.connect(self._exit_replay)
        lay.addWidget(btn_exit)
        return box

    def _build_bottombar(self):
        box = QtWidgets.QFrame(); box.setObjectName("Card")
        lay = QtWidgets.QHBoxLayout(box); lay.setContentsMargins(10, 6, 10, 6)
        b1 = QtWidgets.QPushButton("Clear Session"); b1.clicked.connect(self._clear)
        lay.addWidget(b1)
        self.chk_console = QtWidgets.QCheckBox("Show raw serial in console"); self.chk_console.setChecked(True)
        lay.addWidget(self.chk_console)
        lay.addStretch(1)
        be = QtWidgets.QPushButton("Export Excel (.xlsx)"); be.setObjectName("Primary")
        be.clicked.connect(self._export_excel); lay.addWidget(be)
        bh = QtWidgets.QPushButton("Export HTML Report")
        bh.clicked.connect(self._export_html); lay.addWidget(bh)
        bx = QtWidgets.QPushButton("Export XML Report"); bx.clicked.connect(self._export_xml)
        lay.addWidget(bx)
        return box

    # -- plot builders --
    def _dual(self, title, ll, lc, rl, rc):
        p = pg.PlotWidget(); p.setBackground(BG_NORMAL)
        p.setTitle(title, color="#ddd", size="10pt")
        p.showGrid(x=True, y=True, alpha=0.2)
        p.getAxis("left").setLabel(ll, color=lc); p.getAxis("left").setPen(lc)
        p.getAxis("left").setTextPen(lc)
        vb2 = pg.ViewBox(); p.showAxis("right"); p.scene().addItem(vb2)
        p.getAxis("right").linkToView(vb2); vb2.setXLink(p)
        p.getAxis("right").setLabel(rl, color=rc); p.getAxis("right").setPen(rc)
        p.getAxis("right").setTextPen(rc)
        # faint dashed = RAW; bold solid = FILTERED (overlay both)
        cl_raw = p.plot(pen=pg.mkPen(lc, width=1, style=Qt.DashLine))
        cr_raw = pg.PlotCurveItem(pen=pg.mkPen(rc, width=1, style=Qt.DashLine)); vb2.addItem(cr_raw)
        cl = p.plot(pen=pg.mkPen(lc, width=2))
        cr = pg.PlotCurveItem(pen=pg.mkPen(rc, width=2)); vb2.addItem(cr)
        p._vb2 = vb2; p._cl = cl; p._cr = cr; p._cl_raw = cl_raw; p._cr_raw = cr_raw
        p.getViewBox().sigResized.connect(
            lambda *_: (vb2.setGeometry(p.getViewBox().sceneBoundingRect()),
                        vb2.linkedViewChanged(p.getViewBox(), vb2.XAxis)))
        return p

    def _single(self, title, ylabel, color):
        p = pg.PlotWidget(); p.setBackground(BG_NORMAL)
        p.setTitle(title, color="#ddd", size="10pt")
        p.showGrid(x=True, y=True, alpha=0.2)
        p.setLabel("left", ylabel, color=color)
        p._curve = p.plot(pen=pg.mkPen(color, width=2),
                          fillLevel=0, brush=pg.mkBrush(255, 138, 101, 60))
        return p

    def _scatter(self):
        p = pg.PlotWidget(); p.setBackground(BG_NORMAL)
        p.setTitle("Distance vs Packet-Loss (per node)", color="#ddd", size="10pt")
        p.showGrid(x=True, y=True, alpha=0.25)
        p.setLabel("bottom", "Distance (m)"); p.setLabel("left", "Packet loss (%)")
        p._scatter = pg.ScatterPlotItem(size=7)
        p.addItem(p._scatter)
        p._legend = p.addLegend()
        return p

    def _bar(self):
        b = QtWidgets.QProgressBar(); b.setRange(0, 100); b.setValue(0); b.setFormat("%p%")
        return b

    # ------------------------------------------------------------- serial
    def _refresh_ports(self):
        self.cmb_port.clear()
        for p in serial.tools.list_ports.comports():
            self.cmb_port.addItem(f"{p.device} - {p.description}", p.device)
        if self.cmb_port.count() == 0:
            self.cmb_port.addItem("No ports", None)

    def _add_port(self):
        port = self.cmb_port.currentData()
        if not port:
            self._log("[ERR] no port selected"); return
        if port in self.workers:
            self._log(f"[INFO] {port} already connected"); return
        w = SerialWorker(port, self.cmb_baud.currentText())
        w.node_tel.connect(self._on_node)
        w.tasks.connect(self._on_tasks)
        w.gw_tel.connect(self._on_gw)
        w.ota.connect(self._on_ota)
        w.sec.connect(self._on_sec)
        w.resp.connect(lambda p, o: self._log(f"[{p}] RESP {json.dumps(o)}"))
        w.raw.connect(self._on_raw)
        w.status.connect(self._on_status)
        self.workers[port] = w
        w.start()
        self._update_portlbl()

    def _disconnect_all(self):
        for w in list(self.workers.values()):
            w.stop()
        self.workers.clear()
        self.gw_port = None
        self._update_portlbl()

    def _update_portlbl(self):
        if self.workers:
            self.lbl_ports.setText(", ".join(self.workers.keys()))
        else:
            self.lbl_ports.setText("none")

    def _on_status(self, port, connected, msg):
        self._log(f"[{port}] {msg}")
        if not connected and port in self.workers:
            try:
                self.workers.pop(port).deleteLater()
            except Exception:                              # noqa: BLE001
                pass
            if self.gw_port == port:
                self.gw_port = None
            self._update_portlbl()

    def _on_raw(self, port, line):
        if self.chk_console.isChecked():
            self._log(f"[{port}] {line}")

    def _on_node(self, port, tel):
        if self.replay_active:
            return                          # ignore live data while replaying
        if self.recording:
            self._rec("node", port, tel)
        if self.frozen:
            return
        self._apply_node(port, tel)

    def _apply_node(self, port, tel):
        src = tel.get("src", "gw")          # "node" = direct, "gw" = relayed
        store = self.nodes if src == "node" else self.nodes_gw
        if src == "gw":
            self.gw_port = port             # this port is the gateway (for commands)
        nid = int(tel.get("node", 1))
        nd = store.get(nid)
        if nd is None:
            nd = NodeData(nid); store[nid] = nd
            self._sync_node_combo()
        nd.port = port
        nd.push(tel)

    def _on_gw(self, port, tel):
        if self.replay_active:
            return
        if self.recording:
            self._rec("gw", port, tel)
        self.gw_port = port
        if self.frozen:
            return
        self._apply_gw(port, tel)

    def _apply_gw(self, port, tel):
        self.gw.push(tel)

    def _on_ota(self, port, obj):
        """OTA progress tu node/gateway: {type:ota, node, state, pct, msg}."""
        nid = int(obj.get("node", 0))
        self.ota_state[nid] = obj
        self._log(f"[OTA] node {nid}: {obj.get('state')} {obj.get('pct')}% {obj.get('msg','')}")

    def _on_sec(self, port, obj):
        """Su kien bao mat: replay_drop (goi that bi loai) / replay_test (HIL nap lai)."""
        self.sec_events.append((time.time(), obj))
        ev = obj.get("event")
        if ev == "replay_test":
            self.replay_result = obj
            self._log(f"[SEC] replay_test: age={obj.get('age_s')}s "
                      f"window={obj.get('window_s')}s -> {obj.get('result','?').upper()}")
        elif ev == "replay_drop":
            self._log(f"[SEC] REPLAY BLOCKED: node {obj.get('node')} "
                      f"lech {abs(int(obj.get('gw_epoch',0))-int(obj.get('pkt_epoch',0)))}s "
                      f"(tong {obj.get('drops')} goi bi chan)")

    def _on_tasks(self, port, obj):
        if self.replay_active:
            return
        if self.recording:
            self._rec("tasks", port, obj)
        if self.frozen:
            return
        self._apply_tasks(port, obj)

    def _apply_tasks(self, port, obj):
        nid = int(obj.get("node", 1))
        nd = self.nodes.get(nid)
        if nd is None:
            nd = NodeData(nid); self.nodes[nid] = nd; self._sync_node_combo()
        nd.tasks = obj.get("tasks", [])
        nd.core0 = int(obj.get("c0_cpu", nd.core0))
        nd.core1 = int(obj.get("c1_cpu", nd.core1))

    # ---- recording ----
    def _rec(self, kind, port, data):
        t = time.time()
        if self.record_t0 is None:
            self.record_t0 = t
        self.record_frames.append((round(t - self.record_t0, 3), kind, port, data))

    def _toggle_record(self, on):
        if on:
            if self.replay_active:
                self._exit_replay()
            self.record_frames = []; self.record_t0 = None; self.recording = True
            self.btn_rec.setText("⏹ Stop & Save")
            self.btn_rec.setStyleSheet("background:#b33; font-weight:700;")
            self._log("[REC] recording started")
        else:
            self.recording = False
            self.btn_rec.setText("⏺ Record")
            self.btn_rec.setStyleSheet("")
            self._save_record()

    def _save_record(self):
        if not self.record_frames:
            self._log("[REC] nothing recorded"); return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Save session log",
            f"EdgeProfiler_session_{datetime.datetime.now():%Y%m%d_%H%M%S}.eplog",
            "EdgeProfiler log (*.eplog)")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                meta = {"app": APP_NAME, "version": APP_VERSION,
                        "saved": datetime.datetime.now().isoformat(timespec="seconds"),
                        "frames": len(self.record_frames),
                        "duration_s": self.record_frames[-1][0]}
                f.write("#META " + json.dumps(meta) + "\n")
                for (t, kind, port, data) in self.record_frames:
                    f.write(json.dumps({"t": t, "kind": kind, "port": port, "data": data},
                                       separators=(",", ":")) + "\n")
            self._log(f"[REC] saved {len(self.record_frames)} frames -> {path}")
        except Exception as exc:                           # noqa: BLE001
            self._log(f"[REC] save failed: {exc}")

    # ---- drag & drop a log file ----
    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()

    def dropEvent(self, e):
        for url in e.mimeData().urls():
            p = url.toLocalFile()
            if p.lower().endswith((".eplog", ".jsonl", ".ndjson", ".log")):
                self._load_log(p); return
        self._log("[REPLAY] drop a .eplog session file to replay")

    def _open_log(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Open session log", "", "EdgeProfiler log (*.eplog *.jsonl *.ndjson *.log)")
        if path:
            self._load_log(path)

    def _load_log(self, path):
        frames = []
        try:
            with open(path, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    try:
                        o = json.loads(line)
                    except (json.JSONDecodeError, ValueError):
                        continue
                    if "t" in o and "kind" in o and "data" in o:
                        frames.append((float(o["t"]), o["kind"], o.get("port", ""), o["data"]))
        except Exception as exc:                           # noqa: BLE001
            self._log(f"[REPLAY] load failed: {exc}"); return
        if not frames:
            self._log("[REPLAY] no frames found in file"); return
        frames.sort(key=lambda x: x[0])
        self.replay_frames = frames
        self.replay_dur = frames[-1][0]
        self._enter_replay(path)

    # ---- replay ----
    def _enter_replay(self, path):
        if self.recording:
            self.btn_rec.setChecked(False)   # stop & save current recording first
        self.replay_active = True
        self.replay_paused = True
        self.replay_time = 0.0
        self.replay_pos = 0
        self.btn_rec.setEnabled(False)
        self.replay_bar.setVisible(True)
        self.btn_play.setText("▶ Play")
        self._replay_rebuild(0.0)
        self._update_replay_ui()
        self._log(f"[REPLAY] loaded {len(self.replay_frames)} frames "
                  f"({self.replay_dur:.1f}s). Drag the slider, Play, zoom charts as if live.")

    def _exit_replay(self):
        self.replay_active = False
        self.replay_paused = True
        self.replay_bar.setVisible(False)
        self.btn_rec.setEnabled(True)
        self.nodes.clear(); self.nodes_gw.clear(); self.gw = GatewayData()
        self._sync_node_combo()
        self._log("[REPLAY] exited — live data resumes")

    def _feed_frame(self, frame):
        _, kind, port, data = frame
        if kind == "node":
            self._apply_node(port, data)
        elif kind == "gw":
            self._apply_gw(port, data)
        elif kind == "tasks":
            self._apply_tasks(port, data)

    def _replay_feed_to(self, t):
        fr = self.replay_frames
        while self.replay_pos < len(fr) and fr[self.replay_pos][0] <= t:
            self._feed_frame(fr[self.replay_pos])
            self.replay_pos += 1

    def _replay_rebuild(self, t):
        self.nodes.clear(); self.nodes_gw.clear(); self.gw = GatewayData()
        self.replay_pos = 0
        self._replay_feed_to(t)
        self._sync_node_combo()

    def _replay_tick(self):
        if not self.replay_active or self.replay_paused:
            return
        self.replay_time = min(self.replay_dur, self.replay_time + 0.1 * self.replay_speed)
        self._replay_feed_to(self.replay_time)
        self._update_replay_ui()
        if self.replay_time >= self.replay_dur:
            self.replay_paused = True
            self.btn_play.setText("▶ Play")

    def _replay_play_pause(self):
        if not self.replay_active:
            return
        if self.replay_time >= self.replay_dur:
            self._replay_rebuild(0.0); self.replay_time = 0.0
        self.replay_paused = not self.replay_paused
        self.btn_play.setText("▶ Play" if self.replay_paused else "⏸ Pause")

    def _replay_seek(self, val):
        if not self.replay_active or self.replay_dur <= 0:
            return
        t = self.replay_dur * val / 1000.0
        if t < self.replay_time:
            self._replay_rebuild(t)
        else:
            self._replay_feed_to(t)
        self.replay_time = t
        self.lbl_time.setText(f"{self._mmss(t)} / {self._mmss(self.replay_dur)}")

    def _replay_speed_change(self, text):
        self.replay_speed = {"0.5x": 0.5, "1x": 1.0, "2x": 2.0,
                             "4x": 4.0, "10x": 10.0, "Max": 60.0}.get(text, 1.0)

    @staticmethod
    def _mmss(s):
        s = int(s)
        return f"{s // 60:02d}:{s % 60:02d}"

    def _update_replay_ui(self):
        if self.replay_dur > 0:
            self.slider.blockSignals(True)
            self.slider.setValue(int(1000 * self.replay_time / self.replay_dur))
            self.slider.blockSignals(False)
        self.lbl_time.setText(f"{self._mmss(self.replay_time)} / {self._mmss(self.replay_dur)}")

    def _on_source_change(self, text):
        self.source_mode = text
        self._sync_node_combo()
        self._refresh_views()

    def _active_nodes(self):
        """Dict feeding the main plots/tiles, per the Source selector."""
        if self.source_mode.startswith("Gateway"):
            return self.nodes_gw
        return self.nodes                   # Node direct (also primary for Both)

    def _gw_worker(self):
        if self.gw_port and self.gw_port in self.workers:
            return self.workers[self.gw_port]
        for w in self.workers.values():
            if w.is_gateway:
                return w
        return next(iter(self.workers.values()), None)

    # ------------------------------------------------------------- views
    def _sync_node_combo(self):
        cur = self.cmb_node.currentData()
        ids = sorted(set(self.nodes.keys()) | set(self.nodes_gw.keys()))
        self.cmb_node.blockSignals(True)
        self.cmb_node.clear()
        for nid in ids:
            self.cmb_node.addItem(f"Node {nid}", nid)
        if cur is not None:
            i = self.cmb_node.findData(cur)
            if i >= 0:
                self.cmb_node.setCurrentIndex(i)
        self.cmb_node.blockSignals(False)

    def _cur_node(self):
        nid = self.cmb_node.currentData()
        return self._active_nodes().get(nid)

    def _state_snapshot(self):
        """State for the node-centric test sequencer (reads node-DIRECT data)."""
        nid = self.cmb_node.currentData()
        nd = self.nodes.get(nid) or next(iter(self.nodes.values()), None)
        st = {"gw_age": (time.time() - self.gw.last_seen) if self.gw.last_seen else 99,
              "has_node": nd is not None}
        if nd:
            def lastv(k):
                v = nd.s[k][-1] if nd.s[k] else None
                return None if (isinstance(v, float) and math.isnan(v)) else v
            sen = nd.last.get("sensor", {})
            st.update({
                "online": nd.online, "fsm": nd.fsm, "alert": nd.alert, "algo": nd.algo,
                "rtt": lastv("rtt"), "pph": lastv("pph"), "enc": lastv("enc"),
                "proc": lastv("proc"), "rssi": lastv("rssi"), "heap": lastv("heap"),
                "t": sen.get("t"), "h": sen.get("h"), "press": sen.get("press"),
                "rx": nd.rx, "sent": nd.sent, "cpu": nd.cpu0,
                "buf_count": nd.buf_count, "buf_stored": nd.buf_stored,
                "buf_flushed": nd.buf_flushed, "buf_dropped": nd.buf_dropped,
                "wdt_s": nd.wdt_s, "reset": nd.reset_reason,
                "boot": nd.boot, "wdt_resets": nd.wdt_resets,
                "sleep_mode": nd.sleep_mode, "sleep_ms": nd.sleep_ms,
                "awake_ms": nd.awake_ms,
            })
        # OTA + replay state cho cac kich ban bao mat/OTA
        node_id = nd.node_id if nd else nid
        st["node_id"] = node_id
        st["ota"] = self.ota_state.get(node_id) or next(iter(self.ota_state.values()), None)
        st["replay"] = self.replay_result
        st["has_gw"] = self._gw_worker() is not None
        # Du lieu node do GATEWAY chuyen tiep (OTA tu xa: chi can cam gateway, node o tren cao)
        gw_nd = self.nodes_gw.get(node_id) or next(iter(self.nodes_gw.values()), None)
        if gw_nd:
            st["gw_node_id"] = gw_nd.node_id
            st["gw_fw_ver"] = gw_nd.fw_ver
            st["gw_online"] = gw_nd.online
        return st

    def _refresh_views(self):
        # gateway
        if self.gw.x:
            xs = list(self.gw.x)
            self.c_heap.setData(xs, list(self.gw.heap))
            self.bar_c0.setValue(int(self.gw.c0[-1]) if self.gw.c0 else 0)
            self.bar_c1.setValue(int(self.gw.c1[-1]) if self.gw.c1 else 0)
            g = self.gw.last
            age = time.time() - self.gw.last_seen if self.gw.last_seen else 999
            self.lbl_gw.setText(
                f"heap {int(self.gw.heap[-1]):,} B | up {g.get('uptime',0)}s | "
                f"nodes {g.get('nodes_online',0)} | wifi {'OK' if g.get('wifi') else 'no'} | "
                f"firebase {'OK' if g.get('fb_ok') else 'FAIL'} | tel {age:.0f}s ago")

        # overview table
        self._refresh_table()

        # selected node plots+tiles
        nd = self._cur_node()
        if nd and nd.x:
            xs = list(nd.x)
            def s(k): return list(nd.s[k])
            show_raw = self.plot_mode in ("Both (raw+filtered)", "Raw only")
            show_filt = self.plot_mode in ("Both (raw+filtered)", "Filtered only")

            def setc(curve, ys, show):
                curve.setData(xs, ys) if show else curve.setData([], [])

            # Climate / Environment: overlay RAW (dashed) + FILTERED (solid)
            setc(self.p_climate._cl, s("tf"), show_filt)
            setc(self.p_climate._cl_raw, s("t"), show_raw)
            setc(self.p_climate._cr, s("hf"), show_filt)
            setc(self.p_climate._cr_raw, s("h"), show_raw)
            setc(self.p_env._cl, s("pressf"), show_filt)
            setc(self.p_env._cl_raw, s("press"), show_raw)
            setc(self.p_env._cr, s("luxf"), show_filt)
            setc(self.p_env._cr_raw, s("lux"), show_raw)
            # Network / Performance / Power: raw only
            self.p_net._cl.setData(xs, s("rssi")); self.p_net._cr.setData(xs, s("snr"))
            self.p_perf._cl.setData(xs, s("rtt")); self.p_perf._cr.setData(xs, s("dist"))
            self.p_power._cl.setData(xs, s("awake")); self.p_power._cr.setData(xs, s("sleep"))
            bg = BG_ALERT if nd.alert else BG_NORMAL
            for p in (self.p_climate, self.p_env, self.p_net, self.p_perf):
                p.setBackground(bg)
            # tiles ("—" when offline / value not available)
            def last(k): return nd.s[k][-1] if nd.s[k] else 0
            def fmt(k, spec):
                v = last(k)
                if isinstance(v, float) and math.isnan(v):
                    return "—"
                return spec.format(v)
            self.tiles["rtt"].set_value(fmt("rtt", "{:.0f}"))
            self.tiles["dist"].set_value(fmt("dist", "{:.1f}"))
            self.tiles["pdr"].set_value(fmt("pdr", "{:.0f}"))
            self.tiles["pph"].set_value(fmt("pph", "{:.0f}"))
            self.tiles["proc"].set_value(fmt("proc", "{:.0f}"))
            self.tiles["enc"].set_value(fmt("enc", "{:.0f}"))
            self.tiles["rssi"].set_value(fmt("rssi", "{:.0f}"))
            hv = last("heap")
            self.tiles["heap"].set_value(f"{int(hv):,}" if hv else "—")
            self.tiles["cpu"].set_value(f"{nd.cpu0}%")
            self.tiles["sleep"].set_value(f"{nd.sleep_mode} {nd.sleep_ms}ms")
            wdt_txt = f"{nd.reset_reason}·#{nd.boot}"
            if nd.wdt_resets:
                wdt_txt += f" ⚠{nd.wdt_resets}"
            self.tiles["wdt"].set_value(wdt_txt)
            self.lbl_algo.setText(nd.algo)
            self.lbl_filter.setText(nd.filter_algo)
            if not nd.online:
                self.lbl_fsm.setText("OFFLINE (no LoRa)")
                self.lbl_fsm.setObjectName("FsmEmerg")
                self.lbl_fsm.style().unpolish(self.lbl_fsm); self.lbl_fsm.style().polish(self.lbl_fsm)
            else:
                self._set_fsm(nd.fsm, nd.alert)
            # throughput / counters — all computed BY THE NODE
            self.lbl_rate.setText(f"{nd.sps:.2f} sample/s (node)")
            self.lbl_rx.setText(f"Delivered {nd.rx} / Sent {nd.sent}")
            deliv = (100.0 * nd.rx / nd.sent) if nd.sent else 0.0
            self.lbl_deliv.setText(f"Success {deliv:.0f}%")

            # offline buffer plot + status
            if nd.buf_series:
                bx = list(range(len(nd.buf_series)))
                self.p_buf._curve.setData(bx, list(nd.buf_series))
            cap = nd.buf_cap or 1
            self.lbl_buf_use.setText(f"{nd.buf_count} / {nd.buf_cap} records")
            free_pct = 100.0 * (cap - nd.buf_count) / cap
            self.lbl_buf_free.setText(f"{free_pct:.0f}% free")
            self.lbl_buf_bytes.setText(f"{nd.buf_count * RECORD_BYTES:,} B used")
            self.lbl_buf_tot.setText(
                f"stored {nd.buf_stored} · flushed {nd.buf_flushed} · dropped {nd.buf_dropped}")
            self.buf_bar.setValue(int(100 * nd.buf_count / cap))
            # event log (last 25)
            evs = list(nd.events)[-25:]
            txt = "\n".join(datetime.datetime.fromtimestamp(t).strftime("%H:%M:%S") + "  " + m
                            for t, m in evs)
            if txt != getattr(self, "_last_off_txt", None):
                self.off_log.setPlainText(txt)
                self.off_log.verticalScrollBar().setValue(self.off_log.verticalScrollBar().maximum())
                self._last_off_txt = txt

        # global counters across all nodes (from the active source)
        act = self._active_nodes()
        tot_rx = sum(n.rx for n in act.values())
        tot_sent = sum(n.sent for n in act.values())
        tot_deliv = (100.0 * tot_rx / tot_sent) if tot_sent else 0.0
        self.lbl_tot.setText(f"All nodes: Delivered {tot_rx} / Sent {tot_sent} ({tot_deliv:.0f}%)")

        # Both: compare node-direct vs gateway-relayed (did the gateway get it right?)
        if self.source_mode.startswith("Both"):
            nid = self.cmb_node.currentData()
            nd_d, nd_g = self.nodes.get(nid), self.nodes_gw.get(nid)
            if nd_d and nd_g:
                td = nd_d.last.get("sensor", {}).get("t")
                tg = nd_g.last.get("sensor", {}).get("t")
                ok = (isinstance(td, (int, float)) and isinstance(tg, (int, float))
                      and abs(td - tg) < 0.6)
                self.lbl_compare.setText(
                    f"COMPARE node↔gw: T {td}/{tg} · node delivered {nd_d.rx} · "
                    f"gw received {nd_g.rx}  {'✓ gateway OK' if ok else '✗ differ/stale'}")
            elif nd_d and not nd_g:
                self.lbl_compare.setText("COMPARE: gateway NOT receiving this node (relay empty)")
            else:
                self.lbl_compare.setText("")
        else:
            self.lbl_compare.setText("")

        self._refresh_corr()

    def _refresh_table(self):
        act = self._active_nodes()
        ids = sorted(act.keys())
        self.table.setRowCount(len(ids))

        def cell(nd, k, spec):
            v = nd.s[k][-1] if nd.s[k] else None
            if v is None or (isinstance(v, float) and math.isnan(v)):
                return "—"
            return spec.format(v)

        for r, nid in enumerate(ids):
            nd = act[nid]
            sen = nd.last.get("sensor", {})
            linkup = bool(nd.online)
            fsm = nd.fsm if linkup else "OFFLINE"
            vals = [str(nid), f"{sen.get('t', 0):.1f}", f"{sen.get('h', 0):.0f}",
                    cell(nd, "rssi", "{:.0f}"), cell(nd, "pdr", "{:.0f}"),
                    cell(nd, "rtt", "{:.0f}"), cell(nd, "dist", "{:.1f}"),
                    fsm, "●" if linkup else "○"]
            for c, v in enumerate(vals):
                it = QtWidgets.QTableWidgetItem(v)
                if c == 7:
                    col = {"SAFE": "#26c281", "WARN": "#ffd166",
                           "EMERGENCY": "#ff5c5c", "OFFLINE": "#ff5c5c"}.get(fsm, "#ccc")
                    it.setForeground(QtGui.QColor(col))
                if c == 8:
                    it.setForeground(QtGui.QColor("#26c281" if linkup else "#ff5c5c"))
                self.table.setItem(r, c, it)

    def _refresh_corr(self):
        spots = []
        act = self._active_nodes()
        for idx, nid in enumerate(sorted(act.keys())):
            nd = act[nid]
            col = NODE_COLORS[idx % len(NODE_COLORS)]
            for (d, loss) in list(nd.corr):
                spots.append({"pos": (d, loss), "brush": pg.mkBrush(col), "size": 7, "pen": None})
        self.p_corr._scatter.setData(spots)

    def _set_fsm(self, fsm, alert):
        self.lbl_fsm.setText(fsm)
        name = "FsmSafe"
        if fsm == "EMERGENCY":
            name = "FsmEmerg"
        elif fsm == "WARN" or alert:
            name = "FsmWarn"
        self.lbl_fsm.setObjectName(name)
        self.lbl_fsm.style().unpolish(self.lbl_fsm); self.lbl_fsm.style().polish(self.lbl_fsm)

    def _open_detail(self, key):
        nd = self._cur_node()
        if not nd:
            return
        if key == "wdt":
            QtWidgets.QMessageBox.information(
                self, f"Watchdog — Node {nd.node_id}",
                f"Watchdog timeout: {nd.wdt_s} s\n"
                f"Last reset reason: {nd.reset_reason}\n"
                f"Boot count (RTC-retained): {nd.boot}\n"
                f"Times reset BY watchdog: {nd.wdt_resets}\n\n"
                f"Test: chập chân GPIO0 của node xuống GND để giả lập treo — "
                f"watchdog sẽ reset node, reason đổi thành TASK_WDT và boot count tăng.")
            return
        titles = {"rtt": "Response Time (ms)", "dist": "Distance (m)", "pdr": "PDR (%)",
                  "pph": "Bandwidth (pkt/h)", "proc": "Processing time (µs)",
                  "enc": "Encryption time (µs)", "rssi": "RSSI (dBm)", "heap": "Node heap (B)",
                  "cpu": "Node CPU (%)", "sleep": "Sleep duration per cycle (ms)"}
        colors = {"rtt": C["rtt"], "dist": C["dist"], "pdr": C["pdr"], "pph": C["pph"],
                  "proc": C["proc"], "enc": C["enc"], "rssi": C["rssi"], "heap": C["heap"],
                  "cpu": "#ff7043", "sleep": "#80cbc4"}
        d = DetailDialog(nd, key, titles.get(key, key), colors.get(key, "#4fc3f7"), self)
        d.show()
        self.detail_dialogs.append(d)

    def _toggle_freeze(self, on):
        self.frozen = on
        self.btn_freeze.setText("▶ Resume" if on else "⏸ Freeze")

    def _calibrate_distance(self):
        import math
        nd = self._cur_node()
        if not nd or not nd.s["rssi"]:
            self._log("[CAL] no node/RSSI to calibrate from"); return
        rssi = nd.s["rssi"][-1]
        d = self.spin_cal_d.value()
        n = self.spin_n.value()
        # ref(@1m) = RSSI(d) + 10*n*log10(d)
        ref = rssi + 10.0 * n * math.log10(max(d, 0.01))
        CALIB["ref"] = ref
        self.spin_ref.blockSignals(True); self.spin_ref.setValue(ref); self.spin_ref.blockSignals(False)
        self._log(f"[CAL] node {nd.node_id}: RSSI {rssi:.0f} dBm @ {d} m  ->  RSSI@1m = {ref:.1f} dBm (n={n})")

    def _open_help(self):
        HelpDialog(self).show()

    def _open_rtos(self):
        d = RtosDialog(lambda: self.nodes.get(self.cmb_node.currentData()), self)
        d.show()
        self.detail_dialogs.append(d)

    # ------------------------------------------------------------- tests
    def _run_test(self):
        if self.seq and self.seq.isRunning():
            self._log("[WARN] a test is already running"); return
        scn = self.cmb_scn.currentText()
        gw_scn = scn.startswith("OTA") or scn.startswith("LoRa Replay")
        if gw_scn:
            if self._gw_worker() is None:
                self._log("[ERR] cam cong GATEWAY truoc (kich ban nay chay tren gateway)"); return
        elif not self.nodes:
            self._log("[ERR] connect the NODE's USB port first (tests evaluate the node)"); return
        extra = {"ota_url": self.ed_ota_url.text()}
        self.seq = TestSequencer(self._gw_worker, self._state_snapshot, scn, extra)
        self.seq.log.connect(self._log)
        self.seq.result.connect(self._on_result)
        self.seq.done.connect(lambda: (self.btn_run.setEnabled(True), self.btn_run.setText("RUN")))
        self.btn_run.setEnabled(False); self.btn_run.setText("RUNNING…")
        self.seq.start()

    def _on_result(self, res):
        self.test_results.append(res)
        # verification accuracy = fraction of node checks that passed
        self.anomaly_injected += 1
        if res["passed"]:
            self.anomaly_detected += 1

    # ------------------------------------------------------------- misc
    def _log(self, text):
        self.console.appendPlainText(text)
        sb = self.console.verticalScrollBar(); sb.setValue(sb.maximum())

    def _clear(self):
        self.nodes.clear(); self.nodes_gw.clear(); self.gw = GatewayData()
        self.test_results.clear(); self.anomaly_injected = self.anomaly_detected = 0
        self.session_start = datetime.datetime.now()
        self.console.clear(); self.table.setRowCount(0); self._sync_node_combo()
        self._log("[INFO] session cleared")

    # ------------------------------------------------------------- report
    def _node_stats(self, nd, key):
        v = [x for x in nd.acc.get(key, [])
             if isinstance(x, (int, float)) and not math.isnan(x)]
        if not v:
            return {"min": None, "max": None, "avg": None}
        return {"min": round(min(v), 2), "max": round(max(v), 2), "avg": round(statistics.mean(v), 2)}

    def _report_ctx(self):
        end = datetime.datetime.now()
        passed = sum(1 for r in self.test_results if r["passed"])
        acc = (100.0 * self.anomaly_detected / self.anomaly_injected) if self.anomaly_injected else None
        nodes = []
        for nid in sorted(self.nodes.keys()):
            nd = self.nodes[nid]
            nodes.append({
                "id": nid, "fsm": nd.fsm, "algo": nd.algo,
                "t": self._node_stats(nd, "t"), "h": self._node_stats(nd, "h"),
                "lux": self._node_stats(nd, "lux"), "press": self._node_stats(nd, "press"),
                "rssi": self._node_stats(nd, "rssi"), "snr": self._node_stats(nd, "snr"),
                "rtt": self._node_stats(nd, "rtt"), "dist": self._node_stats(nd, "dist"),
                "pdr": self._node_stats(nd, "pdr"), "pph": self._node_stats(nd, "pph"),
                "proc": self._node_stats(nd, "proc"), "enc": self._node_stats(nd, "enc"),
                "samples": len(nd.acc.get("t", [])),
                "buf_cap": nd.buf_cap, "buf_count": nd.buf_count,
                "buf_stored": nd.buf_stored, "buf_flushed": nd.buf_flushed,
                "buf_dropped": nd.buf_dropped, "backlog_recs": len(nd.backlog_log),
            })
        gw = self.gw.last
        return {
            "app": APP_NAME, "version": APP_VERSION,
            "start": self.session_start.strftime("%Y-%m-%d %H:%M:%S"),
            "end": end.strftime("%Y-%m-%d %H:%M:%S"),
            "duration": str(end - self.session_start).split(".")[0],
            "tests": self.test_results, "passed": passed,
            "failed": len(self.test_results) - passed, "total": len(self.test_results),
            "anomaly_acc": round(acc, 1) if acc is not None else None,
            "nodes": nodes, "ports": list(self.workers.keys()),
            "gw_heap_start": int(self.gw.heap[0]) if self.gw.heap else None,
            "gw_heap_end": int(self.gw.heap[-1]) if self.gw.heap else None,
            "gw": gw,
        }

    def _export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            self._log("[XLSX] openpyxl missing — run: pip install openpyxl"); return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Export session to Excel",
            f"EdgeProfiler_{datetime.datetime.now():%Y%m%d_%H%M%S}.xlsx", "Excel (*.xlsx)")
        if not path:
            return

        def clean(v):
            if isinstance(v, float) and math.isnan(v):
                return None
            return v

        wb = Workbook()
        ws = wb.active; ws.title = "Summary"
        passed = sum(1 for r in self.test_results if r["passed"])
        rows = [
            ["EdgeProfiler session report"],
            ["Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Session start", self.session_start.strftime("%Y-%m-%d %H:%M:%S")],
            ["Nodes", ", ".join(map(str, sorted(self.nodes.keys())))],
            ["Ports", ", ".join(self.workers.keys())],
            ["Tests passed", f"{passed}/{len(self.test_results)}"],
        ]
        for r in rows:
            ws.append(r)
        ws["A1"].font = Font(bold=True, size=14)

        # one sheet per node with the full time series
        for nid, nd in sorted(self.nodes.items()):
            s = wb.create_sheet(f"Node{nid}")
            header = ["sample"] + NodeData.SERIES
            s.append(header)
            for c in range(1, len(header) + 1):
                s.cell(row=1, column=c).font = Font(bold=True)
            xs = list(nd.x)
            series = {k: list(nd.s[k]) for k in NodeData.SERIES}
            for i in range(len(xs)):
                row = [xs[i]]
                for k in NodeData.SERIES:
                    row.append(clean(series[k][i]) if i < len(series[k]) else None)
                s.append(row)

        # gateway resources
        if self.gw.x:
            s = wb.create_sheet("Gateway")
            s.append(["sample", "heap", "c0_cpu", "c1_cpu"])
            gx = list(self.gw.x); h = list(self.gw.heap); c0 = list(self.gw.c0); c1 = list(self.gw.c1)
            for i in range(len(gx)):
                s.append([gx[i], h[i] if i < len(h) else None,
                          c0[i] if i < len(c0) else None, c1[i] if i < len(c1) else None])

        # events (offline / watchdog / backlog)
        s = wb.create_sheet("Events")
        s.append(["node", "time", "event"])
        for nid, nd in sorted(self.nodes.items()):
            for (t, msg) in nd.events:
                s.append([nid, datetime.datetime.fromtimestamp(t).strftime("%H:%M:%S"), msg])

        # tests
        s = wb.create_sheet("Tests")
        s.append(["test", "result", "detail", "timestamp"])
        for r in self.test_results:
            s.append([r["name"], "PASS" if r["passed"] else "FAIL", r["detail"], r["ts"]])

        try:
            wb.save(path)
            self._log(f"[XLSX] exported -> {path}")
        except Exception as exc:                           # noqa: BLE001
            self._log(f"[XLSX] save failed: {exc}")

    def _export_html(self):
        ctx = self._report_ctx()
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Save HTML", f"EdgeProfiler_{datetime.datetime.now():%Y%m%d_%H%M%S}.html",
            "HTML (*.html)")
        if not path:
            return
        with open(path, "w", encoding="utf-8") as f:
            f.write(Template(HTML_TEMPLATE).render(**ctx))
        self._log(f"[INFO] HTML report saved: {path}")

    def _export_xml(self):
        ctx = self._report_ctx()
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Save XML", f"EdgeProfiler_{datetime.datetime.now():%Y%m%d_%H%M%S}.xml",
            "XML (*.xml)")
        if not path:
            return
        root = ET.Element("EdgeProfilerReport", app=APP_NAME, version=APP_VERSION)
        s = ET.SubElement(root, "Session")
        ET.SubElement(s, "Start").text = ctx["start"]
        ET.SubElement(s, "End").text = ctx["end"]
        ET.SubElement(s, "Duration").text = ctx["duration"]
        ET.SubElement(s, "Ports").text = ", ".join(ctx["ports"])
        if ctx["anomaly_acc"] is not None:
            ET.SubElement(s, "AnomalyDetectionAccuracy").text = str(ctx["anomaly_acc"])
        ts = ET.SubElement(root, "testsuite", name="HIL", tests=str(ctx["total"]),
                           failures=str(ctx["failed"]))
        for r in ctx["tests"]:
            tc = ET.SubElement(ts, "testcase", name=r["name"])
            if not r["passed"]:
                ET.SubElement(tc, "failure", message=r["detail"]).text = r["detail"]
            else:
                ET.SubElement(tc, "system-out").text = r["detail"]
        ns = ET.SubElement(root, "Nodes")
        for n in ctx["nodes"]:
            e = ET.SubElement(ns, "Node", id=str(n["id"]), fsm=n["fsm"], algo=n["algo"])
            for key in ("t", "h", "lux", "press", "rssi", "snr", "rtt", "dist", "pdr", "pph"):
                st = n[key]
                ET.SubElement(e, key, min=str(st["min"]), max=str(st["max"]), avg=str(st["avg"]))
            ET.SubElement(e, "OfflineBuffer", capacity=str(n["buf_cap"]),
                          current=str(n["buf_count"]), stored=str(n["buf_stored"]),
                          flushed=str(n["buf_flushed"]), dropped=str(n["buf_dropped"]),
                          backlog_replayed=str(n["backlog_recs"]))
        ET.indent(root, space="  ")
        ET.ElementTree(root).write(path, encoding="utf-8", xml_declaration=True)
        self._log(f"[INFO] XML report saved: {path}")

    def closeEvent(self, ev):
        try:
            if self.seq and self.seq.isRunning():
                self.seq.abort(); self.seq.wait(800)
            for w in list(self.workers.values()):
                w.stop()
        finally:
            ev.accept()


# ==========================================================================
HTML_TEMPLATE = r"""<!DOCTYPE html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>{{app}} Report</title>
<style>
:root{--bg:#0f1419;--card:#1b212b;--ink:#e6e6e6;--mut:#8aa;--ok:#26c281;--bad:#ff5c5c;--acc:#4fc3f7;}
*{box-sizing:border-box}body{margin:0;font-family:Segoe UI,Arial,sans-serif;background:var(--bg);color:var(--ink);padding:24px}
h1{margin:0 0 4px}.sub{color:var(--mut);margin-bottom:18px}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:14px;margin-bottom:16px}
.card{background:var(--card);border:1px solid #2a3340;border-radius:12px;padding:16px}
.card h2{margin:0 0 12px;font-size:13px;letter-spacing:.5px;color:var(--acc);text-transform:uppercase}
table{width:100%;border-collapse:collapse}td,th{padding:7px 6px;border-bottom:1px solid #2a3340;text-align:left;font-size:13px}
th{color:var(--mut)}.big{font-size:28px;font-weight:700}.mut{color:var(--mut)}
.pill{padding:3px 10px;border-radius:999px;font-weight:700;font-size:12px}
.pass{background:rgba(38,194,129,.15);color:var(--ok)}.fail{background:rgba(255,92,92,.15);color:var(--bad)}
.ok{color:var(--ok)}.bad{color:var(--bad)}
</style></head><body>
<h1>{{app}} &mdash; HIL Multi-Node Report</h1>
<div class="sub">ESP32-S3 LoRa Edge System &nbsp;|&nbsp; v{{version}} &nbsp;|&nbsp; ports: {{ports|join(', ')}}</div>
<div class="kpis">
 <div class="card"><div class="mut">Tests Passed</div><div class="big ok">{{passed}}/{{total}}</div></div>
 <div class="card"><div class="mut">Tests Failed</div><div class="big {{'bad' if failed else 'ok'}}">{{failed}}</div></div>
 <div class="card"><div class="mut">Anomaly Detection</div><div class="big">{{ anomaly_acc if anomaly_acc is not none else 'N/A' }}{{ '%' if anomaly_acc is not none else '' }}</div></div>
 <div class="card"><div class="mut">Nodes Seen</div><div class="big">{{nodes|length}}</div></div>
 <div class="card"><div class="mut">Duration</div><div class="big">{{duration}}</div></div>
</div>
<div class="card"><h2>Session</h2><table>
 <tr><th>Start</th><td>{{start}}</td><th>End</th><td>{{end}}</td></tr>
 <tr><th>Gateway heap start</th><td>{{ '{:,}'.format(gw_heap_start) if gw_heap_start else 'N/A' }} B</td>
     <th>Gateway heap end</th><td>{{ '{:,}'.format(gw_heap_end) if gw_heap_end else 'N/A' }} B</td></tr>
</table></div>
<div class="card" style="margin-top:16px"><h2>Test Pass/Fail Matrix</h2><table>
 <tr><th>#</th><th>Test</th><th>Result</th><th>Detail</th><th>Time</th></tr>
 {% for r in tests %}<tr><td>{{loop.index}}</td><td>{{r.name}}</td>
 <td><span class="pill {{'pass' if r.passed else 'fail'}}">{{'PASS' if r.passed else 'FAIL'}}</span></td>
 <td class="mut">{{r.detail}}</td><td class="mut">{{r.ts}}</td></tr>{% endfor %}
 {% if not tests %}<tr><td colspan="5" class="mut">No tests executed.</td></tr>{% endif %}
</table></div>
{% for n in nodes %}
<div class="card" style="margin-top:16px"><h2>Node {{n.id}} &mdash; {{n.algo}} &mdash; FSM {{n.fsm}}</h2>
<table><tr><th>Metric</th><th>Min</th><th>Max</th><th>Avg</th></tr>
 <tr><td>Temperature (°C)</td><td>{{n.t.min}}</td><td>{{n.t.max}}</td><td>{{n.t.avg}}</td></tr>
 <tr><td>Humidity (%)</td><td>{{n.h.min}}</td><td>{{n.h.max}}</td><td>{{n.h.avg}}</td></tr>
 <tr><td>Light (lux)</td><td>{{n.lux.min}}</td><td>{{n.lux.max}}</td><td>{{n.lux.avg}}</td></tr>
 <tr><td>Pressure (hPa)</td><td>{{n.press.min}}</td><td>{{n.press.max}}</td><td>{{n.press.avg}}</td></tr>
 <tr><td>RSSI (dBm)</td><td>{{n.rssi.min}}</td><td>{{n.rssi.max}}</td><td>{{n.rssi.avg}}</td></tr>
 <tr><td>SNR (dB)</td><td>{{n.snr.min}}</td><td>{{n.snr.max}}</td><td>{{n.snr.avg}}</td></tr>
 <tr><td>Response RTT (ms)</td><td>{{n.rtt.min}}</td><td>{{n.rtt.max}}</td><td>{{n.rtt.avg}}</td></tr>
 <tr><td>Distance (m)</td><td>{{n.dist.min}}</td><td>{{n.dist.max}}</td><td>{{n.dist.avg}}</td></tr>
 <tr><td>PDR (%)</td><td>{{n.pdr.min}}</td><td>{{n.pdr.max}}</td><td>{{n.pdr.avg}}</td></tr>
 <tr><td>Bandwidth (pkt/h)</td><td>{{n.pph.min}}</td><td>{{n.pph.max}}</td><td>{{n.pph.avg}}</td></tr>
 <tr><td>Processing (µs)</td><td>{{n.proc.min}}</td><td>{{n.proc.max}}</td><td>{{n.proc.avg}}</td></tr>
 <tr><td>Encryption (µs)</td><td>{{n.enc.min}}</td><td>{{n.enc.max}}</td><td>{{n.enc.avg}}</td></tr>
</table>
<p class="mut" style="margin-top:8px">Offline store-and-forward: buffer {{n.buf_count}}/{{n.buf_cap}} now ·
stored {{n.buf_stored}} · flushed {{n.buf_flushed}} · dropped {{n.buf_dropped}} ·
backlog records replayed: {{n.backlog_recs}}
{% if n.buf_count == 0 %}<span class="ok">(buffer empty — all data delivered)</span>{% else %}<span class="bad">(buffer not empty)</span>{% endif %}</p>
</div>
{% endfor %}
<p class="sub" style="margin-top:18px">Generated by {{app}} v{{version}} — Automated HIL Testing &amp; Monitoring.</p>
</body></html>
"""

DARK_QSS = """
QWidget{background:#0f1419;color:#e6e6e6;font-family:'Segoe UI',Arial;font-size:13px;}
QFrame#Card{background:#1b212b;border:1px solid #2a3340;border-radius:10px;}
QFrame#Tile{background:#11161d;border:1px solid #2a3340;border-radius:8px;}
QFrame#Tile:hover{border:1px solid #4fc3f7;}
QLabel#Header{color:#4fc3f7;font-weight:700;letter-spacing:1px;padding:2px 0 6px 0;}
QLabel#FsmSafe{background:#13351f;color:#26c281;border-radius:6px;padding:6px;font-weight:700;}
QLabel#FsmWarn{background:#3a2e10;color:#ffd166;border-radius:6px;padding:6px;font-weight:700;}
QLabel#FsmEmerg{background:#3a1414;color:#ff5c5c;border-radius:6px;padding:6px;font-weight:700;}
QPushButton{background:#263041;border:1px solid #34465e;border-radius:6px;padding:6px 12px;}
QPushButton:hover{background:#30405a;}
QPushButton:checked{background:#7a4;}
QPushButton#Primary{background:#1f6feb;border:none;font-weight:700;}
QPushButton#Primary:hover{background:#388bfd;}
QPushButton:disabled{background:#20262f;color:#667;}
QComboBox{background:#11161d;border:1px solid #34465e;border-radius:6px;padding:4px 8px;}
QComboBox QAbstractItemView{background:#11161d;selection-background-color:#1f6feb;}
QPlainTextEdit#Console{background:#0a0d11;color:#c8e1ff;border:1px solid #2a3340;border-radius:6px;
 font-family:'Consolas','Courier New',monospace;font-size:12px;}
QProgressBar{background:#11161d;border:1px solid #34465e;border-radius:6px;text-align:center;height:18px;}
QProgressBar::chunk{background:#1f6feb;border-radius:5px;}
QTableWidget{background:#11161d;gridline-color:#2a3340;border:none;}
QHeaderView::section{background:#1b212b;color:#8aa;border:none;padding:4px;}
QTabBar::tab{background:#1b212b;color:#aab;padding:6px 12px;border-top-left-radius:6px;border-top-right-radius:6px;}
QTabBar::tab:selected{background:#263041;color:#fff;}
QTabWidget::pane{border:1px solid #2a3340;border-radius:6px;}
"""


def main():
    pg.setConfigOptions(antialias=True)
    app = QtWidgets.QApplication(sys.argv)
    app.setStyleSheet(DARK_QSS)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
